import json
import os
import tempfile
import unittest

import api_client
import store_profiles
from openpyxl import Workbook, load_workbook

from models import Product
from processor import init_output_workbook, write_product_row
from store_profiles import (
    DEFAULT_PROFILE_FILE,
    category_fields,
    deterministic_suggestion,
    find_store,
    import_template_profile,
    load_store_profiles,
    match_store_category,
    read_source_context,
    resolve_fields,
    validate_profile,
    validate_store_profile,
)


ROOT = os.path.dirname(os.path.abspath(__file__))
STORE_ID = "demo_store"


def sample_profiles():
    return store_profiles.normalize_profiles({
        "version": 1,
        "global_fields": {"B": "옥션/G마켓", "N": "무제한"},
        "stores": [{
            "id": STORE_ID,
            "name": "演示店铺",
            "aliases": ["DEMO"],
            "fields": {
                "C": "demo_account", "D": "demo_account",
                "AE": "100001", "AF": "100002", "AG": "100003",
                "AH": "-100", "AI": "-100",
            },
            "categories": [
                {"id": "bags", "name": "包类",
                 "fields": {"AL": "3", "AM": "200001"}},
                {"id": "home", "name": "家居收纳类",
                 "fields": {"AL": "35", "AM": "200002"}},
            ],
        }],
    })


class StoreProfileTests(unittest.TestCase):
    @staticmethod
    def category_id(data, store_id, name):
        store = find_store(data, store_id)
        return next(category["id"] for category in store["categories"]
                    if category["name"] == name)

    def test_profile_fields_are_resolved_from_explicit_user_data(self):
        data = sample_profiles()
        bags_id = self.category_id(data, STORE_ID, "包类")
        home_id = self.category_id(data, STORE_ID, "家居收纳类")
        bags = resolve_fields({}, STORE_ID, bags_id, data)
        home = resolve_fields({}, STORE_ID, home_id, data)

        self.assertEqual(bags["C"], "demo_account")
        self.assertEqual((bags["AL"], bags["AM"]), ("3", "200001"))
        self.assertEqual((home["AL"], home["AM"]), ("35", "200002"))
        self.assertEqual(validate_profile(STORE_ID, bags_id, data), [])
        self.assertEqual(validate_store_profile(STORE_ID, data), [])

    def test_defaults_are_loaded_when_user_profile_does_not_exist(self):
        original_profile = store_profiles.PROFILE_FILE
        original_defaults = store_profiles.DEFAULT_PROFILE_FILE
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                store_profiles.PROFILE_FILE = os.path.join(temp_dir, "store_profiles.yaml")
                store_profiles.DEFAULT_PROFILE_FILE = DEFAULT_PROFILE_FILE
                data = store_profiles.load_store_profiles()
        finally:
            store_profiles.PROFILE_FILE = original_profile
            store_profiles.DEFAULT_PROFILE_FILE = original_defaults

        self.assertEqual(data["stores"], [])
        self.assertEqual(data["global_fields"]["N"], "무제한")

    def test_universal_template_uses_unlimited_listing_period(self):
        template = os.path.join(ROOT, "uploader", "韩国上传模板.xlsx")
        workbook = load_workbook(template, read_only=True, data_only=False)
        try:
            self.assertEqual(workbook["NEW 일반상품"]["N8"].value, "무제한")
            for column in ("C", "D", "AE", "AF", "AG", "AH", "AI", "AL", "AM"):
                self.assertIsNone(workbook["NEW 일반상품"][f"{column}8"].value)
        finally:
            workbook.close()

    def test_profile_values_override_legacy_template_values(self):
        data = sample_profiles()
        bags_id = self.category_id(data, STORE_ID, "包类")
        resolved = resolve_fields(
            {"B": "legacy", "C": "legacy_store", "AL": "999", "AM": "999"},
            STORE_ID, bags_id, data)
        self.assertEqual(resolved["B"], "옥션/G마켓")
        self.assertEqual(resolved["C"], "demo_account")
        self.assertEqual(resolved["AL"], "3")
        self.assertEqual(resolved["AM"], "200001")

    def test_filename_aliases_select_store_and_category_without_ai(self):
        data = sample_profiles()
        result = deterministic_suggestion(
            r"D:\采集表\DEMO-家居收纳类-20260720.xlsx", [], data)
        self.assertEqual(result["store_id"], STORE_ID)
        self.assertEqual(
            category_fields(STORE_ID, result["category_id"], data)["AL"], "35")
        self.assertEqual(result["confidence"], 1.0)

    def test_mixed_products_resolve_different_al_am_profiles(self):
        data = sample_profiles()
        home = Product(title="厨房抽屉收纳盒", tag="家居 收纳")
        bag = Product(title="女士通勤手提包", tag="箱包 包袋")

        home_match = match_store_category(
            STORE_ID, home, category_zh="家居收纳类", data=data, use_ai=False)
        bag_match = match_store_category(
            STORE_ID, bag, category_zh="包类", data=data, use_ai=False)

        self.assertEqual(category_fields(STORE_ID, home_match["category_id"], data)["AL"], "35")
        self.assertEqual(category_fields(STORE_ID, bag_match["category_id"], data)["AL"], "3")
        self.assertEqual(bag_match["method"], "rule")

    def test_deepseek_category_name_is_resolved_to_internal_id(self):
        data = sample_profiles()
        original = api_client.deepseek_chat
        api_client.deepseek_chat = lambda *args, **kwargs: (
            '{"category_id":"包类","confidence":0.2,"evidence":["最接近的大类"]}')
        try:
            result = match_store_category(
                STORE_ID, Product(title="无法直接关键词匹配"), data=data, use_ai=True)
        finally:
            api_client.deepseek_chat = original

        self.assertEqual(result["category_id"], self.category_id(data, STORE_ID, "包类"))
        self.assertEqual(result["confidence"], 0.2)

    def test_deepseek_empty_choice_is_retried_as_forced_broad_category(self):
        data = sample_profiles()
        home_name = next(
            category["name"] for category in find_store(data, STORE_ID)["categories"]
            if category["name"] != "包类")
        responses = iter([
            '{"category_id":"","confidence":0.1,"evidence":["没有精确类目"]}',
            json.dumps({"category_id": home_name, "confidence": 0.4,
                        "evidence": ["最接近家居大类"]}, ensure_ascii=False),
        ])
        prompts = []
        original = api_client.deepseek_chat

        def fake_deepseek(prompt, **kwargs):
            prompts.append(prompt)
            return next(responses)

        api_client.deepseek_chat = fake_deepseek
        try:
            result = match_store_category(
                STORE_ID, Product(title="客厅沙发边桌"),
                category_path="사이드테이블", category_zh="边桌",
                data=data, use_ai=True)
        finally:
            api_client.deepseek_chat = original

        self.assertEqual(result["category_id"], self.category_id(data, STORE_ID, home_name))
        self.assertEqual(result["deepseek_calls"], 2)
        self.assertIn("必须", prompts[0])
        self.assertIn("禁止返回空值", prompts[1])

    def test_category_profiles_only_keep_name_as_match_reference(self):
        data = sample_profiles()
        for store in data["stores"]:
            for category in store["categories"]:
                self.assertNotIn("aliases", category)
                self.assertNotIn("match_hints", category)

    def test_old_template_import_splits_store_and_category_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "T77上架表格-宠物类.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "NEW 일반상품"
            values = {"C": "store77", "D": "store77", "AE": "1001", "AF": "1002",
                      "AG": "1003", "AH": "-202", "AI": "-202",
                      "AL": "9", "AM": "90001"}
            for column, value in values.items():
                sheet[f"{column}8"] = value
            workbook.save(path)

            data, store_id, category_id = import_template_profile(path, {"stores": []})
            resolved = resolve_fields({}, store_id, category_id, data)
            self.assertEqual(resolved["C"], "store77")
            self.assertEqual(resolved["AE"], "1001")
            self.assertEqual(resolved["AL"], "9")
            self.assertEqual(resolved["AM"], "90001")

    def test_source_context_reads_headers_and_analysis_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "collector.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "采集数据"
            headers = [f"字段{i}" for i in range(13)]
            headers[0], headers[2], headers[4], headers[5], headers[11] = (
                "ParentSKU", "标题", "标签", "品牌", "来源链接")
            row = [""] * 13
            row[0], row[2], row[4], row[5], row[11] = (
                "P1", "收纳盒", "家居收纳", "品牌A", "https://example.com/item")
            sheet.append(headers)
            sheet.append(row)
            workbook.save(path)

            context = read_source_context(path)

            self.assertEqual(context["sheet_name"], "采集数据")
            self.assertIn("ParentSKU", context["headers"])
            self.assertEqual(context["rows"][0]["title"], "收纳盒")
            self.assertEqual(context["rows"][0]["brand"], "品牌A")

    def test_writer_uses_store_and_category_fields_with_dynamic_values(self):
        data = sample_profiles()
        template = os.path.join(ROOT, "uploader", "韩国上传模板.xlsx")
        _, sheet, template_fixed = init_output_workbook(
            template, "unused.xlsx", cfg={})
        bags_id = self.category_id(data, STORE_ID, "包类")
        fixed = resolve_fields(template_fixed, STORE_ID, "", data)
        product = Product(parent_sku="SKU-1", ai_title="테스트 상품")
        product.result.update({
            "K": "esm-code", "L": "a-code", "M": "g-code",
            "O": 12345, "W": "단독형", "X": "선택형", "Y": "",
            "Z": "https://example.com/main.jpg", "AA": "", "AB": "detail",
        })
        write_product_row(
            sheet, 0, product, fixed, 8,
            category_fields(STORE_ID, bags_id, data))

        self.assertEqual(sheet["C8"].value, "demo_account")
        self.assertEqual(sheet["AE8"].value, "100001")
        self.assertEqual(sheet["AL8"].value, "3")
        self.assertEqual(sheet["AM8"].value, "200001")
        self.assertEqual(sheet["K8"].value, "esm-code")
        self.assertEqual(sheet["L8"].value, "a-code")
        self.assertEqual(sheet["M8"].value, "g-code")

    def test_writer_supports_different_al_am_in_one_output(self):
        data = sample_profiles()
        template = os.path.join(ROOT, "uploader", "韩国上传模板.xlsx")
        _, sheet, template_fixed = init_output_workbook(
            template, "unused.xlsx", cfg={})
        fixed = resolve_fields(template_fixed, STORE_ID, "", data)
        home_id = self.category_id(data, STORE_ID, "家居收纳类")
        bags_id = self.category_id(data, STORE_ID, "包类")
        products = [Product(ai_title="home"), Product(ai_title="bag")]
        for product in products:
            product.result.update({
                "K": "1", "L": "2", "M": "3", "O": 100,
                "W": "x", "X": "y", "Y": "", "Z": "z", "AA": "", "AB": "a",
            })

        write_product_row(
            sheet, 0, products[0], fixed, 8,
            category_fields(STORE_ID, home_id, data))
        write_product_row(
            sheet, 1, products[1], fixed, 8,
            category_fields(STORE_ID, bags_id, data))

        self.assertEqual((sheet["AL8"].value, sheet["AM8"].value), ("35", "200002"))
        self.assertEqual((sheet["AL9"].value, sheet["AM9"].value), ("3", "200001"))


if __name__ == "__main__":
    unittest.main()

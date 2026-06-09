# Author: Administrator
# Created: 2026-05-25
"""GUI for 耀我科技上传器 v2.0 — main window, panels, dialogs, test panel."""

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import ttkbootstrap as tb
import os
import threading
import time

from models import ProductStatus
from config_manager import load_config, save_config, load_prompts, save_prompts

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))

# ---- Color scheme & constants ----
STATUS_COLORS = {
    ProductStatus.PENDING: "#95a5a6",
    ProductStatus.PHASE1_TITLE: "#f39c12",
    ProductStatus.PHASE1_CATEGORY: "#f39c12",
    ProductStatus.PHASE1_PRICE: "#f39c12",
    ProductStatus.PHASE1_DONE: "#3498db",
    ProductStatus.PHASE2_MAIN_IMG: "#2980b9",
    ProductStatus.PHASE2_VARIANT: "#2980b9",
    ProductStatus.PHASE2_DETAIL: "#2980b9",
    ProductStatus.DONE: "#27ae60",
    ProductStatus.FAILED: "#c0392b",
}


class UploaderApp:
    """Main application window for 耀我科技上传器 v2.0."""

    def __init__(self, startup_warnings=None, restore_state=None):
        self.warnings = startup_warnings or []
        self.root = tb.Window(themename="cosmo")
        self.root.title("耀我科技上传器 v2.0")
        self.root.geometry("1100x700")
        self._setup_menu()
        self._setup_toolbar()
        self._setup_main_area()
        self._setup_bottom_bar()
        self._apply_warnings()
        # State
        self.source_path = None
        self.template_path = None
        self.products = []
        self._processing = False
        self._paused = False
        self._stopped = False
        self._pause_event = threading.Event()
        self._pause_event.set()
        # Stats
        self._stats = {"deepseek": 0, "haomingai": 0, "upload": 0}
        # Restore saved state
        if restore_state:
            self._restore_state(restore_state)

    # ============================================================
    # WAVE 4 — task-05: Menu + Toolbar + Main Layout + Bottom Bar
    # ============================================================

    def _setup_menu(self):
        bar = tk.Menu(self.root)
        # File menu
        fm = tk.Menu(bar, tearoff=0)
        fm.add_command(label="选择采集表...", command=self._select_source)
        fm.add_command(label="导入更多SKU...", command=self._import_more_skus)
        fm.add_command(label="选择模板...", command=self._select_template)
        fm.add_separator()
        fm.add_command(label="导出上架表...", command=self._do_export)
        fm.add_separator()
        fm.add_command(label="退出", command=self._on_close)
        bar.add_cascade(label="文件", menu=fm)
        # Process menu
        pm = tk.Menu(bar, tearoff=0)
        pm.add_command(label="开始处理", command=self._start_processing)
        pm.add_command(label="暂停", command=self._toggle_pause)
        pm.add_command(label="停止", command=self._stop)
        pm.add_separator()
        pm.add_command(label="重试全部失败", command=self._retry_all_failed)
        pm.add_separator()
        pm.add_command(label="筛选失败产品", command=self._filter_failed)
        pm.add_command(label="显示全部", command=self._refresh_list)
        pm.add_separator()
        pm.add_command(label="勾选所有失败", command=self._check_all_failed)
        pm.add_command(label="取消全部勾选", command=self._uncheck_all)
        pm.add_separator()
        pm.add_command(label="清空列表", command=self._clear_list)
        bar.add_cascade(label="处理", menu=pm)
        # Settings menu
        sm = tk.Menu(bar, tearoff=0)
        sm.add_command(label="API 配置...", command=self._open_api_settings)
        sm.add_command(label="价格系数...", command=self._open_price_settings)
        sm.add_command(label="提示词管理...", command=self._open_prompt_manager)
        sm.add_separator()
        sm.add_command(label="处理设置...", command=self._open_batch_settings)
        # Theme submenu
        theme_menu = tk.Menu(sm, tearoff=0)
        for t in ["cosmo", "flatly", "journal", "litera", "lumen", "minty",
                   "pulse", "sandstone", "united", "yeti", "morph",
                   "cyborg", "darkly", "solar", "superhero", "vapor"]:
            theme_menu.add_command(label=t, command=lambda tn=t: self._set_theme(tn))
        sm.add_cascade(label="主题切换", menu=theme_menu)
        bar.add_cascade(label="设置", menu=sm)
        # Help menu
        hm = tk.Menu(bar, tearoff=0)
        hm.add_command(label="使用说明", command=self._open_help)
        bar.add_cascade(label="帮助", menu=hm)
        self.root.config(menu=bar)

    def _setup_toolbar(self):
        tbar = tb.Frame(self.root, padding=4)
        tbar.pack(fill=tk.X)
        tb.Label(tbar, text="采集表:").pack(side=tk.LEFT, padx=2)
        self._src_btn = tb.Button(tbar, text="无", width=20, command=self._select_source,
                                  bootstyle="outline-secondary")
        self._src_btn.pack(side=tk.LEFT, padx=2)
        tb.Label(tbar, text="  模板:").pack(side=tk.LEFT, padx=2)
        self._tpl_btn = tb.Button(tbar, text="无", width=20, command=self._select_template,
                                  bootstyle="outline-secondary")
        self._tpl_btn.pack(side=tk.LEFT, padx=2)
        tb.Separator(tbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        tb.Label(tbar, text="标题:").pack(side=tk.LEFT, padx=2)
        self._title_profile_var = tk.StringVar(value="title")
        self._title_profile_cb = tb.Combobox(tbar, textvariable=self._title_profile_var,
                                              values=["title"], state="readonly", width=10)
        self._title_profile_cb.pack(side=tk.LEFT, padx=2)
        tb.Label(tbar, text="生图:").pack(side=tk.LEFT, padx=(8, 2))
        all_profiles = list(load_prompts().keys())
        img_profiles = [k for k in all_profiles if k != "title"] or ["generic"]
        self._img_profile_var = tk.StringVar(value=img_profiles[0])
        self._img_profile_cb = tb.Combobox(tbar, textvariable=self._img_profile_var,
                                            values=img_profiles, state="readonly", width=14)
        self._img_profile_cb.pack(side=tk.LEFT, padx=2)
        # Image API selector
        cfg = load_config()
        current_api = cfg.get("image_api", "haomingai")
        self._img_api_var = tk.StringVar(value=current_api)
        self._img_api_cb = tb.Combobox(tbar, textvariable=self._img_api_var,
                                         values=["haomingai", "hfsyapi"],
                                         state="readonly", width=10)
        self._img_api_cb.pack(side=tk.LEFT, padx=2)
        self._img_api_cb.bind("<<ComboboxSelected>>", self._on_img_api_changed)
        # Platform selector for collector table format
        tb.Label(tbar, text="  平台:").pack(side=tk.LEFT, padx=(8,2))
        self._platform_var = tk.StringVar(value="shein")
        self._platform_cb = tb.Combobox(tbar, textvariable=self._platform_var,
                                          values=["shein", "aliexpress"],
                                          state="readonly", width=10)
        self._platform_cb.pack(side=tk.LEFT, padx=2)
        self._platform_cb.bind("<<ComboboxSelected>>", self._on_platform_changed)
        tb.Separator(tbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=6)
        self._start_btn = tb.Button(tbar, text="开 始", command=self._start_processing,
                                    bootstyle="success")
        self._start_btn.pack(side=tk.LEFT, padx=4)
        self._export_btn = tb.Button(tbar, text="导出上架表", command=self._do_export,
                                     bootstyle="primary")
        self._export_btn.pack(side=tk.LEFT, padx=4)
        self._batch_regen_btn = tb.Button(tbar, text="重新生成选中", command=self._regenerate_checked,
                                          bootstyle="warning")
        self._batch_regen_btn.pack(side=tk.LEFT, padx=4)
        tb.Label(tbar, text="", font=("", 1)).pack(side=tk.LEFT, padx=6)
        self._mode_var = tk.StringVar(value="等待开始...")
        tb.Label(tbar, textvariable=self._mode_var, foreground="gray").pack(side=tk.LEFT)

    def _setup_main_area(self):
        pw = tb.Panedwindow(self.root, orient=tk.HORIZONTAL)
        pw.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)
        # Left: product list
        left = tb.Frame(pw)
        pw.add(left, weight=60)
        self._setup_product_list(left)
        # Right: notebook (preview / test)
        right = tb.Frame(pw)
        pw.add(right, weight=40)
        self._setup_right_panel(right)

    def _setup_bottom_bar(self):
        bf = tb.Frame(self.root, padding=4)
        bf.pack(fill=tk.X)
        self._stat_label = tb.Label(bf, text="DeepSeek: 0 | 生图: 0 | 上传: 0", font=("", 8))
        self._stat_label.pack(side=tk.LEFT, padx=4)
        self._prog = tb.Progressbar(bf, mode="determinate", length=300)
        self._prog.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
        self._prog_label = tb.Label(bf, text="0/0", font=("", 8))
        self._prog_label.pack(side=tk.LEFT, padx=4)
        self._pause_btn = tb.Button(bf, text="暂 停", command=self._toggle_pause,
                                    bootstyle="warning-outline")
        self._pause_btn.pack(side=tk.RIGHT, padx=4)
        self._stop_btn = tb.Button(bf, text="停 止", command=self._stop,
                                   bootstyle="danger-outline")
        self._stop_btn.pack(side=tk.RIGHT, padx=4)

    def _apply_warnings(self):
        if self.warnings:
            msg = "\n".join(self.warnings)
            self.root.after(500, lambda: messagebox.showwarning("启动提醒", msg))

    def _restore_state(self, state):
        """Restore uploader state from a saved session."""
        self.source_path = state.get("source_path", "")
        self.template_path = state.get("template_path", "")
        self.products = state.get("products", [])
        # Update toolbar buttons
        if self.source_path:
            self._src_btn.configure(text=os.path.basename(self.source_path))
        if self.template_path:
            self._tpl_btn.configure(text=os.path.basename(self.template_path))
        # Restore profile settings
        img_p = state.get("img_profile", "generic")
        title_p = state.get("title_profile", "title")
        if img_p in self._img_profile_cb["values"]:
            self._img_profile_var.set(img_p)
        if title_p in self._title_profile_cb["values"]:
            self._title_profile_var.set(title_p)
        # Restore stats
        self._stats = state.get("stats", {"deepseek": 0, "haomingai": 0, "upload": 0})
        self._stat_label.configure(
            text=f"DeepSeek: {self._stats['deepseek']} | 生图: {self._stats['haomingai']} | 上传: {self._stats['upload']}")
        # Set output path from saved state
        output = state.get("output_path", "")
        if output:
            from processor import detect_completed
            done_skus = detect_completed(output)
            for prod in self.products:
                if prod.parent_sku in done_skus:
                    prod.status = ProductStatus.DONE
        # Refresh list
        self._refresh_list()
        pending = sum(1 for p in self.products if p.status != ProductStatus.DONE)
        if pending:
            self._mode_var.set(f"已恢复 — {pending} 个待处理")
        else:
            self._mode_var.set("已恢复 — 全部完成")

    # ============================================================
    # WAVE 4 — task-06: Product List (Treeview)
    # ============================================================

    def _setup_product_list(self, parent):
        cols = ("check", "#", "SKU", "status", "category", "title")
        self._tree = tb.Treeview(parent, columns=cols, show="headings",
                                  selectmode="browse")
        self._tree.heading("check", text="☑")
        self._tree.heading("#", text="#")
        self._tree.heading("SKU", text="ParentSKU")
        self._tree.heading("status", text="状态")
        self._tree.heading("category", text="类目(中文)")
        self._tree.heading("title", text="标题")
        self._tree.column("check", width=30, anchor=tk.CENTER)
        self._tree.column("#", width=35, anchor=tk.CENTER)
        self._tree.column("SKU", width=90)
        self._tree.column("status", width=65, anchor=tk.CENTER)
        self._tree.column("category", width=90)
        self._tree.column("title", width=170)
        # Scrollbar
        sb = tb.Scrollbar(parent, orient=tk.VERTICAL, command=self._tree.yview)
        self._tree.configure(yscrollcommand=sb.set)
        self._tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        # Tag configs for status text color only (preserves selection highlight)
        for status, color in STATUS_COLORS.items():
            self._tree.tag_configure(status.name, foreground=color)
        self._tree.bind("<<TreeviewSelect>>", self._on_product_select)
        self._tree.bind("<Double-1>", self._on_product_double_click)
        self._tree.bind("<Button-1>", self._on_tree_click)
        self._tree.bind("<Button-3>", self._on_tree_right_click)
        self._checked_products = set()  # set of tree item IDs

    def _on_product_select(self, event):
        sel = self._tree.selection()
        if not sel:
            return
        idx = int(self._tree.index(sel[0]))
        if 0 <= idx < len(self.products):
            self._show_preview(self.products[idx])

    def _on_product_double_click(self, event):
        """Double-click a product row → open category search to replace."""
        sel = self._tree.selection()
        if not sel:
            return
        idx = int(self._tree.index(sel[0]))
        if not (0 <= idx < len(self.products)):
            return
        self._open_category_search(self.products[idx], idx)

    def _open_category_search(self, prod, idx):
        """Dialog: search categories by Chinese name, click to replace."""
        from config_manager import load_category_zh, load_categories

        dlg = tk.Toplevel(self.root)
        dlg.title("搜索类目 — 双击替换")
        dlg.geometry("550x450")
        dlg.transient(self.root)

        tk.Label(dlg, text="输入中文关键词搜索类目:").pack(pady=6)
        search_var = tk.StringVar()
        search_entry = tb.Entry(dlg, textvariable=search_var, width=50)
        search_entry.pack(fill=tk.X, padx=10, pady=2)
        search_entry.focus()

        # Load categories and Chinese cache
        cat_cache = load_category_zh()
        cat_names = load_categories(
            os.path.join(SKILL_DIR, "uploader", "카테고리목록(类别代码K列).xls"))

        listbox = tk.Listbox(dlg, height=18, width=60)
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=6)

        # Build display list: (path, zh_name, codes)
        all_items = []
        for path, codes in cat_names.items():
            zh = cat_cache.get(path, path)
            all_items.append((path, zh, codes))
        # Sort by Chinese name
        all_items.sort(key=lambda x: x[1])

        display_map = []  # maps listbox row → all_items index

        def _refresh_list(filter_text=""):
            listbox.delete(0, tk.END)
            display_map.clear()
            ft = filter_text.lower()
            for i, (path, zh, codes) in enumerate(all_items):
                if ft in zh.lower() or ft in path.lower():
                    listbox.insert(tk.END, zh)
                    display_map.append(i)

        _refresh_list()
        search_var.trace_add("write", lambda *a: _refresh_list(search_var.get()))

        def _on_select(event=None):
            sel = listbox.curselection()
            if not sel:
                return
            item_idx = display_map[sel[0]]
            chosen_path, chosen_zh, chosen_codes = all_items[item_idx]
            # Update product
            prod.result["K"] = chosen_codes.get("esm_code", "")
            prod.result["L"] = chosen_codes.get("auction", "")
            prod.result["M"] = chosen_codes.get("gmarket", "")
            prod.result["K_path"] = chosen_path
            prod.result["K_zh"] = chosen_zh
            prod.logs.append(f"{time.strftime('%H:%M:%S')} 手动换类目: {chosen_path} ({chosen_zh})")
            self.update_product_status(idx, prod)
            self._show_preview(prod)
            dlg.destroy()

        listbox.bind("<Double-1>", _on_select)

        tb.Button(dlg, text="替换选中类目", command=_on_select).pack(pady=6)
        tk.Label(dlg, text="提示: 输入中文搜索, 双击列表中的类目即可替换",
                 fg="gray").pack(pady=4)

    # ============================================================
    # WAVE 4 — task-07: Preview Panel
    # ============================================================

    def _setup_right_panel(self, parent):
        self._notebook = tb.Notebook(parent)
        self._notebook.pack(fill=tk.BOTH, expand=True)
        # Preview tab
        prev_frame = tb.Frame(self._notebook, padding=8)
        self._notebook.add(prev_frame, text="预览")
        self._setup_preview_content(prev_frame)
        # Test tab (built in Wave 5)
        test_frame = tb.Frame(self._notebook, padding=8)
        self._notebook.add(test_frame, text="测试")
        self._setup_test_content(test_frame)

    def _setup_preview_content(self, parent):
        # Scrollable canvas
        canvas = tk.Canvas(parent, highlightthickness=0)
        scrollbar = tb.Scrollbar(parent, orient=tk.VERTICAL, command=canvas.yview)
        self._preview_inner = tb.Frame(canvas)
        self._preview_inner.bind("<Configure>",
                                 lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._preview_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        # Image pair — use Canvas for better resize control
        img_frame = tb.LabelFrame(self._preview_inner, text="主图")
        img_frame.pack(fill=tk.X, pady=(0, 8))
        self._img_orig = tk.Canvas(img_frame, width=180, height=180, bg="#e0e0e0", relief=tk.SUNKEN)
        self._img_orig.pack(side=tk.LEFT, padx=4, fill=tk.BOTH, expand=True)
        self._img_gen = tk.Canvas(img_frame, width=180, height=180, bg="#e0e0e0", relief=tk.SUNKEN)
        self._img_gen.pack(side=tk.LEFT, padx=4, fill=tk.BOTH, expand=True)
        # Regenerate button
        regen_frame = tb.Frame(self._preview_inner)
        regen_frame.pack(fill=tk.X, pady=(0, 4))
        self._regen_btn = tb.Button(regen_frame, text="重新生成主图", command=self._regenerate_main_image)
        self._regen_btn.pack(side=tk.LEFT)
        self._regen_status = tb.Label(regen_frame, text="", foreground="gray")
        self._regen_status.pack(side=tk.LEFT, padx=8)
        # Fields
        self._pv_fields = {}
        for label, key, ro in [("Parent SKU", "parent_sku", False),
                                ("AI 标题", "ai_title", True),
                                ("类目", "category", True),
                                ("价格", "price", True),
                                ("变种", "variants", True)]:
            f = tb.LabelFrame(self._preview_inner, text=label)
            f.pack(fill=tk.X, pady=2)
            if key == "variants":
                w = tk.Text(f, height=5, width=40, state="normal")
            else:
                w = tb.Entry(f, width=45)
            w.pack(fill=tk.X)
            if ro and key != "variants":
                if isinstance(w, tb.Entry):
                    w.configure(state="readonly")
            self._pv_fields[key] = w
        # Log
        log_frame = tb.LabelFrame(self._preview_inner, text="处理日志")
        log_frame.pack(fill=tk.X, pady=2)
        self._log_text = tk.Text(log_frame, height=6, width=40, bg="#2c3e50",
                                 fg="#27ae60", font=("Consolas", 9))
        self._log_text.pack(fill=tk.X)

    def _show_preview(self, prod):
        """Populate preview panel with product data + image previews."""
        # Parent SKU (selectable for copy)
        self._pv_fields["parent_sku"].configure(state="normal")
        self._pv_fields["parent_sku"].delete(0, tk.END)
        self._pv_fields["parent_sku"].insert(0, prod.parent_sku)
        self._pv_fields["parent_sku"].configure(state="readonly")

        self._pv_fields["ai_title"].configure(state="normal")
        self._pv_fields["ai_title"].delete(0, tk.END)
        self._pv_fields["ai_title"].insert(0, prod.ai_title or prod.title)
        self._pv_fields["ai_title"].configure(state="readonly")

        cat = prod.result.get("K", "")
        esm = prod.result.get("L", "")
        gmk = prod.result.get("M", "")
        self._pv_fields["category"].configure(state="normal")
        self._pv_fields["category"].delete(0, tk.END)
        self._pv_fields["category"].insert(0, f"{cat}  (ESM:{esm}  G:{gmk})")
        self._pv_fields["category"].configure(state="readonly")

        price = prod.result.get("O", prod.price)
        self._pv_fields["price"].configure(state="normal")
        self._pv_fields["price"].delete(0, tk.END)
        self._pv_fields["price"].insert(0, f"₩{price}")
        self._pv_fields["price"].configure(state="readonly")

        # Auto-fill test panel inputs
        self._title_input.delete(0, tk.END)
        self._title_input.insert(0, prod.title)

        # Load images in background
        self._img_orig.delete("all")
        self._img_orig.create_text(90, 90, text="加载中...", fill="#888")
        self._img_gen.delete("all")
        self._img_gen.create_text(90, 90, text="加载中...", fill="#888")
        threading.Thread(target=self._load_preview_images, args=(prod,), daemon=True).start()
        # Also load reference image for test panel
        if prod.main_img:
            threading.Thread(target=self._load_test_ref_image, args=(prod.main_img,), daemon=True).start()

        var_text = prod.result.get("Y", "\n".join(
            f"{c}, {s}, 정상, 노출, 100, 100" for c, s, _ in prod.color_sizes
        ))
        self._pv_fields["variants"].delete("1.0", tk.END)
        self._pv_fields["variants"].insert("1.0", var_text)

        self._log_text.delete("1.0", tk.END)
        self._log_text.insert("1.0", "\n".join(prod.logs[-10:]))

    def _regenerate_main_image(self):
        """Re-generate main image for the selected product."""
        sel = self._tree.selection()
        if not sel:
            return
        idx = int(self._tree.index(sel[0]))
        if not (0 <= idx < len(self.products)):
            return
        prod = self.products[idx]
        self._regen_btn.configure(state="disabled")
        self._regen_status.configure(text="生成中...", foreground="#f39c12")
        threading.Thread(target=self._do_regen, args=(idx, prod), daemon=True).start()

    def _do_regen(self, idx, prod):
        try:
            from api_client import generate_image, create_storage_provider
            from config_manager import load_config, load_prompts

            cfg = load_config()
            prompts = load_prompts()
            prompt_key = self._img_profile_var.get()
            prompt_text = prompts.get(prompt_key, prompts.get("generic", ""))
            storage = create_storage_provider(cfg)

            refs = [prod.main_img] + prod.extra_imgs[:3]
            refs = list(dict.fromkeys(refs))
            refs = [u for u in refs if u and u.startswith("http")]

            img_bytes = generate_image(prompt_text, refs)

            new_url = storage.upload(img_bytes, f"main_{prod.parent_sku}_regen.jpg")
            prod.result["Z"] = new_url
            prod.logs.append(f"{time.strftime('%H:%M:%S')} 手动重新生成主图")

            self.root.after(0, lambda: (self._show_preview(prod),
                                        self._regen_btn.configure(state="normal"),
                                        self._regen_status.configure(text="完成", foreground="#27ae60"),
                                        self.update_product_status(idx, prod)))
        except Exception as e:
            self.root.after(0, lambda: (self._regen_btn.configure(state="normal"),
                                        self._regen_status.configure(text=f"失败: {e}", foreground="#c0392b")))

    def _canvas_show_image(self, canvas, img_data, cache_key):
        """Display image bytes on a Canvas, scaled to fit. Main-thread only."""
        from PIL import Image
        import io as _io
        pil_img = Image.open(_io.BytesIO(img_data))
        canvas.update_idletasks()
        cw = canvas.winfo_width() or 400
        ch = canvas.winfo_height() or 400
        if cw < 10:
            cw = 400
        if ch < 10:
            ch = 400
        scale = min(cw / pil_img.width, ch / pil_img.height, 1.0)
        new_w, new_h = int(pil_img.width * scale), int(pil_img.height * scale)
        if new_w < 20:
            new_w = 200
        if new_h < 20:
            new_h = 200
        pil_img = pil_img.resize((new_w, new_h), Image.LANCZOS)
        buf = _io.BytesIO()
        pil_img.save(buf, format="PPM")
        tk_img = tk.PhotoImage(data=buf.getvalue())
        if not hasattr(self, '_img_cache'):
            self._img_cache = {}
        self._img_cache[cache_key] = tk_img
        canvas.delete("all")
        canvas.create_image(cw // 2, ch // 2, image=tk_img, anchor=tk.CENTER)

    def _load_preview_images(self, prod):
        """Download and display product images in preview panel (background thread)."""
        try:
            from api_client import download_image
            if prod.main_img:
                img_data = download_image(prod.main_img)
                self.root.after(0, lambda: self._canvas_show_image(
                    self._img_orig, img_data, prod.parent_sku + "_orig"))
            else:
                self.root.after(0, lambda: self._img_orig.delete("all") or
                    self._img_orig.create_text(90, 90, text="[无图片]", fill="#888"))

            gen_url = prod.result.get("Z", "")
            if gen_url and gen_url.startswith("http"):
                try:
                    img_data = download_image(gen_url)
                    self.root.after(0, lambda: self._canvas_show_image(
                        self._img_gen, img_data, prod.parent_sku + "_gen"))
                except Exception:
                    self.root.after(0, lambda: self._img_gen.delete("all") or
                        self._img_gen.create_text(90, 90, text="[加载失败]", fill="#888"))
            else:
                self.root.after(0, lambda: self._img_gen.delete("all") or
                    self._img_gen.create_text(90, 90, text="[待生成]", fill="#888"))
        except Exception:
            self.root.after(0, lambda: self._img_orig.delete("all") or
                self._img_orig.create_text(90, 90, text="[加载失败]", fill="#888"))

    def _load_test_ref_image(self, img_url):
        """Pre-load reference image for the test panel."""
        try:
            from api_client import download_image
            img_data = download_image(img_url)
            self.root.after(0, lambda: self._canvas_show_image(
                self._img_ref, img_data, "test_ref"))
        except Exception:
            self.root.after(0, lambda: self._img_ref.delete("all") or
                self._img_ref.create_text(90, 90, text="[加载失败]", fill="#888"))

    # ============================================================
    # WAVE 5 — task-10: Test Panel (built alongside for convenience)
    # ============================================================

    def _setup_test_content(self, parent):
        # Test type radio
        type_frame = tb.Frame(parent)
        type_frame.pack(fill=tk.X, pady=4)
        tb.Label(type_frame, text="测试类型:").pack(side=tk.LEFT)
        self._test_type = tk.StringVar(value="title")
        tb.Radiobutton(type_frame, text="标题", variable=self._test_type,
                        value="title", command=self._toggle_test_type).pack(side=tk.LEFT, padx=4)
        tb.Radiobutton(type_frame, text="生图", variable=self._test_type,
                        value="image", command=self._toggle_test_type).pack(side=tk.LEFT, padx=4)
        # Title test area
        self._title_test_frame = tb.Frame(parent)
        self._title_test_frame.pack(fill=tk.BOTH, expand=True)
        tb.Label(self._title_test_frame, text="提示词 (可编辑):").pack(anchor=tk.W, pady=(4, 0))
        self._title_prompt = tk.Text(self._title_test_frame, height=8, width=40)
        self._title_prompt.pack(fill=tk.X, pady=2)
        tb.Label(self._title_test_frame, text="测试产品标题:").pack(anchor=tk.W, pady=(4, 0))
        self._title_input = tb.Entry(self._title_test_frame, width=45)
        self._title_input.pack(fill=tk.X, pady=2)
        btn_frame = tb.Frame(self._title_test_frame)
        btn_frame.pack(fill=tk.X, pady=6)
        tb.Button(btn_frame, text="测试标题", command=self._test_title).pack(side=tk.LEFT, padx=4)
        tb.Button(btn_frame, text="保存提示词", command=self._save_title_prompt).pack(side=tk.LEFT, padx=4)
        tb.Label(self._title_test_frame, text="测试结果:").pack(anchor=tk.W, pady=(8, 0))
        result_row = tb.Frame(self._title_test_frame)
        result_row.pack(fill=tk.X, pady=2)
        self._title_result_var = tk.StringVar()
        self._title_result = tb.Label(result_row, textvariable=self._title_result_var,
                                        relief=tk.SUNKEN, anchor=tk.W, padding=4)
        self._title_result.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self._title_copy_btn = tb.Button(result_row, text="复制", state="disabled",
                                          command=lambda: (self.root.clipboard_clear(),
                                                           self.root.clipboard_append(self._title_result_var.get()),
                                                           self._title_copy_btn.configure(text="已复制"),
                                                           self.root.after(2000, lambda: self._title_copy_btn.configure(text="复制"))))
        self._title_copy_btn.pack(side=tk.LEFT, padx=4)
        self._title_result.pack(fill=tk.X, pady=2)
        # Image test area — scrollable
        self._img_test_frame = tb.Frame(parent)
        img_canvas = tk.Canvas(self._img_test_frame, highlightthickness=0)
        img_scroll = tb.Scrollbar(self._img_test_frame, orient=tk.VERTICAL, command=img_canvas.yview)
        self._img_test_inner = tb.Frame(img_canvas)
        self._img_test_inner.bind("<Configure>",
                                   lambda e: img_canvas.configure(scrollregion=img_canvas.bbox("all")))
        img_canvas.create_window((0, 0), window=self._img_test_inner, anchor="nw")
        img_canvas.configure(yscrollcommand=img_scroll.set)
        img_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        img_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        tb.Label(self._img_test_inner, text="产品类目:").pack(anchor=tk.W, pady=(4, 0))
        img_prompts = [k for k in load_prompts().keys() if k != "title"]
        self._img_category_var = tk.StringVar()
        self._img_category = tb.Combobox(self._img_test_inner, textvariable=self._img_category_var,
                                          values=img_prompts, width=43)
        self._img_category.pack(fill=tk.X, pady=2)
        self._img_category.bind("<<ComboboxSelected>>", self._on_category_selected)
        self._img_category.bind("<FocusOut>", self._on_category_selected)
        tb.Label(self._img_test_inner, text="提示词 (可编辑):").pack(anchor=tk.W, pady=(4, 0))
        self._img_prompt = tk.Text(self._img_test_inner, height=5, width=40)
        self._img_prompt.pack(fill=tk.X, pady=2)
        tb.Label(self._img_test_inner, text="参考图 (选中产品自动加载):").pack(anchor=tk.W, pady=(4, 0))
        self._img_ref = tk.Canvas(self._img_test_inner, height=220, bg="#e0e0e0", relief=tk.SUNKEN)
        self._img_ref.pack(fill=tk.BOTH, expand=True, pady=2)
        self._img_ref.create_text(100, 110, text="[选择产品]", fill="#888")
        btn_frame2 = tb.Frame(self._img_test_inner)
        btn_frame2.pack(fill=tk.X, pady=6)
        tb.Button(btn_frame2, text="测试生图", command=self._test_image).pack(side=tk.LEFT, padx=4)
        tb.Button(btn_frame2, text="保存提示词", command=self._save_image_prompt).pack(side=tk.LEFT, padx=4)
        self._test_status = tb.Label(btn_frame2, text="", foreground="gray")
        self._test_status.pack(side=tk.LEFT, padx=8)
        self._copy_btn = tb.Button(btn_frame2, text="复制生成结果图片URL", state="disabled",
                                    command=lambda: (self.root.clipboard_clear(),
                                                     self.root.clipboard_append(self._test_url_var.get()),
                                                     self._copy_btn.configure(text="已复制"),
                                                     self.root.after(2000, lambda: self._copy_btn.configure(text="复制生成结果图片URL"))))
        self._copy_btn.pack(side=tk.LEFT, padx=4)
        tb.Label(self._img_test_inner, text="生成结果:").pack(anchor=tk.W, pady=(8, 0))
        self._img_result = tk.Canvas(self._img_test_inner, height=330, bg="#f5f5f5", relief=tk.SUNKEN)
        self._img_result.pack(fill=tk.BOTH, expand=True, pady=2)
        self._img_result.create_text(100, 165, text="[等待生成]", fill="#888")
        self._test_url_var = tk.StringVar()
        self._load_test_prompts()

    def _toggle_test_type(self):
        if self._test_type.get() == "title":
            self._img_test_frame.pack_forget()
            self._title_test_frame.pack(fill=tk.BOTH, expand=True)
        else:
            self._title_test_frame.pack_forget()
            self._img_test_frame.pack(fill=tk.BOTH, expand=True)

    def _load_test_prompts(self):
        prompts = load_prompts()
        self._title_prompt.delete("1.0", tk.END)
        self._title_prompt.insert("1.0", prompts.get("title", ""))
        self._img_prompt.delete("1.0", tk.END)
        self._img_prompt.insert("1.0", prompts.get("generic", ""))

    def _test_title(self):
        prod_title = self._title_input.get().strip()
        if not prod_title:
            messagebox.showinfo("提示", "请先输入测试产品标题")
            return
        prompt = self._title_prompt.get("1.0", tk.END).strip()
        if not prompt:
            messagebox.showinfo("提示", "请先填写标题提示词")
            return
        # Replace variables
        prompt_filled = prompt.replace("{product_title}", prod_title).replace("{banned_words}", "").replace("{color_note}", "")
        try:
            from api_client import deepseek_chat
            result = deepseek_chat(prompt_filled, max_tokens=150, temp=0.7)
            self._title_result_var.set(result)
            self._title_copy_btn.configure(state="normal")
        except Exception as e:
            self._title_result_var.set(f"错误: {e}")
            self._title_copy_btn.configure(state="disabled")

    def _save_title_prompt(self):
        prompt = self._title_prompt.get("1.0", tk.END).strip()
        prompts = load_prompts()
        prompts["title"] = prompt
        save_prompts(prompts)
        messagebox.showinfo("提示", "标题提示词已保存到 prompts.yaml")

    def _on_category_selected(self, event=None):
        """Auto-load prompt when user selects an existing category."""
        name = self._img_category_var.get().strip()
        prompts = load_prompts()
        if name and name in prompts:
            self._img_prompt.delete("1.0", tk.END)
            self._img_prompt.insert("1.0", prompts[name])

    def _save_image_prompt(self):
        category = self._img_category_var.get().strip()
        if not category:
            messagebox.showinfo("提示", "请先输入产品类目名称")
            return
        prompt = self._img_prompt.get("1.0", tk.END).strip()
        prompts = load_prompts()
        prompts[category] = prompt
        save_prompts(prompts)
        self._refresh_prompt_profiles()
        messagebox.showinfo("提示", f"生图提示词 [{category}] 已保存到 prompts.yaml")

    def _refresh_prompt_profiles(self):
        """Refresh all prompt dropdowns from prompts.yaml."""
        prompts = load_prompts()
        img_keys = [k for k in prompts if k != "title"] or ["generic"]
        self._img_profile_cb["values"] = img_keys
        if self._img_profile_var.get() not in img_keys:
            self._img_profile_var.set(img_keys[0])
        self._img_category["values"] = img_keys

    def _test_image(self):
        """Async image generation test."""
        prompt = self._img_prompt.get("1.0", tk.END).strip()
        if not prompt:
            messagebox.showinfo("提示", "请先填写生图提示词")
            return
        sel = self._tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在列表中选择一个产品作为参考图来源")
            return
        idx = int(self._tree.index(sel[0]))
        if not (0 <= idx < len(self.products)):
            return
        prod = self.products[idx]
        # Collect up to 4 reference images: main + up to 3 extra images
        refs = [prod.main_img] + prod.extra_imgs[:3]
        refs = list(dict.fromkeys(refs))  # dedup, keep order
        refs = [u for u in refs if u and u.startswith("http")]

        self._test_status.configure(text="生图中...", foreground="#f39c12")
        self._img_result.delete("all")
        self._img_result.create_text(100, 90, text=f"[{len(refs)}张参考图...]", fill="#888")
        threading.Thread(target=self._do_test_image, args=(prompt, refs), daemon=True).start()

    def _do_test_image(self, prompt, ref_urls):
        try:
            from api_client import generate_image, create_storage_provider

            cfg = load_config()
            img_result = generate_image(prompt, ref_urls)

            storage = create_storage_provider(cfg)
            result_url = storage.upload(img_result, "test_output.jpg")
            self.root.after(0, self._show_test_result, img_result, result_url)
        except Exception as e:
            self.root.after(0, self._show_test_error, str(e))

    def _show_test_result(self, img_result, result_url):
        self._canvas_show_image(self._img_result, img_result, "test_gen")
        self._test_status.configure(text="生图成功", foreground="#27ae60")
        self._test_url_var.set(result_url)
        self._copy_btn.configure(state="normal")

    def _show_test_error(self, msg):
        self._img_result.delete("all")
        self._img_result.create_text(100, 100, text=f"错误: {msg[:60]}", fill="#c00")
        self._test_status.configure(text="生图失败", foreground="#c0392b")
        self._test_url_var.set("")
        self._copy_btn.configure(state="disabled")

    # ============================================================
    # WAVE 4 — task-08: Settings Dialogs
    # ============================================================

    def _on_img_api_changed(self, event=None):
        """Save image API choice to config when dropdown changes."""
        cfg = load_config()
        cfg["image_api"] = self._img_api_var.get()
        save_config(cfg)

    def _set_theme(self, name):
        """Switch ttkbootstrap theme."""
        self.root.style.theme_use(name)

    def _open_api_settings(self):
        """API key configuration dialog."""
        dlg = tk.Toplevel(self.root)
        dlg.title("API 配置")
        dlg.geometry("450x460")
        dlg.transient(self.root)
        cfg = load_config()
        storage = cfg.get("storage", {})
        fields = [
            ("DeepSeek Key (必填):", "deepseek_key", cfg.get("deepseek_key", ""), False),
            ("DeepSeek URL:", "deepseek_url", cfg.get("deepseek_url", ""), False),
            ("haomingai Key:", "haomingai_key", cfg.get("haomingai_key", ""), False),
            ("haomingai URL:", "haomingai_url", cfg.get("haomingai_url", ""), False),
            ("hfsyapi Key:", "hfsyapi_key", cfg.get("hfsyapi_key", ""), False),
            ("hfsyapi URL:", "hfsyapi_url", cfg.get("hfsyapi_url", ""), False),
            ("Cloudinary Cloud Name:", "cloud_name", storage.get("cloud_name", ""), False),
            ("Cloudinary API Key:", "api_key", storage.get("api_key", ""), True),
            ("Cloudinary API Secret:", "api_secret", storage.get("api_secret", ""), True),
            ("代理地址 (留空=自动):", "proxy", cfg.get("proxy", ""), False),
        ]
        entries = {}
        for i, (label, key, val, secret) in enumerate(fields):
            tb.Label(dlg, text=label).grid(row=i, column=0, sticky=tk.E, padx=4, pady=3)
            e = tb.Entry(dlg, width=35, show="*" if secret else "")
            if val:
                e.insert(0, str(val))
            e.grid(row=i, column=1, sticky=tk.W, padx=4, pady=3)
            entries[key] = e

        def save():
            for key, e in entries.items():
                v = e.get().strip()
                if key in ("cloud_name", "api_key", "api_secret"):
                    # Save to storage sub-dict
                    if v:
                        cfg.setdefault("storage", {})[key] = v
                        cfg["storage"]["provider"] = "cloudinary"
                    else:
                        cfg.setdefault("storage", {}).pop(key, None)
                else:
                    if v:
                        cfg[key] = v
                    elif key in cfg:
                        del cfg[key]
            save_config(cfg)
            dlg.destroy()
            messagebox.showinfo("提示", "API 配置已保存")

        tb.Button(dlg, text="保存", command=save).grid(row=len(fields), column=0,
                                                          columnspan=2, pady=12)

    def _open_price_settings(self):
        """Price coefficient table editor."""
        dlg = tk.Toplevel(self.root)
        dlg.title("价格系数设置")
        dlg.geometry("400x320")
        dlg.transient(self.root)
        cfg = load_config()
        coeffs = cfg.get("price_coefficients", [[0, 10000, 1.8], [10000, 20000, 1.8],
                                                [20000, 30000, 1.5], [30000, 50000, 1.3],
                                                [50000, 999999999, 1.2]])
        multiplier = cfg.get("global_multiplier", 1.0)

        tb.Label(dlg, text="价格段 (KRW)       系数", font=("", 10, "bold")).pack(pady=4)
        rows_frame = tb.Frame(dlg)
        rows_frame.pack(fill=tk.X, padx=20)
        row_entries = []

        def _add_row(from_val="", to_val="", coeff_val=""):
            rf = tb.Frame(rows_frame)
            rf.pack(fill=tk.X, pady=2)
            e1 = tb.Entry(rf, width=10)
            e1.insert(0, str(from_val))
            e1.pack(side=tk.LEFT, padx=2)
            tb.Label(rf, text="-").pack(side=tk.LEFT)
            e2 = tb.Entry(rf, width=10)
            e2.insert(0, str(to_val))
            e2.pack(side=tk.LEFT, padx=2)
            tb.Label(rf, text="×").pack(side=tk.LEFT)
            e3 = tb.Entry(rf, width=8)
            e3.insert(0, str(coeff_val))
            e3.pack(side=tk.LEFT, padx=2)
            row_entries.append((e1, e2, e3))

        for fr, to, co in coeffs:
            _add_row(fr, to, co)

        def _add_new():
            _add_row()

        tb.Button(dlg, text="+ 新增价格段", command=_add_new).pack(pady=4)

        mult_frame = tb.Frame(dlg)
        mult_frame.pack(pady=6)
        tb.Label(mult_frame, text="全局乘数:").pack(side=tk.LEFT)
        mult_var = tk.StringVar(value=str(multiplier))
        mult_entry = tb.Entry(mult_frame, width=8, textvariable=mult_var)
        mult_entry.pack(side=tk.LEFT, padx=4)

        def save():
            new_coeffs = []
            for e1, e2, e3 in row_entries:
                try:
                    f = float(e1.get())
                    t = float(e2.get())
                    c = float(e3.get())
                    new_coeffs.append([f, t, c])
                except ValueError:
                    pass
            if new_coeffs:
                cfg["price_coefficients"] = new_coeffs
            cfg["global_multiplier"] = float(mult_var.get() or 1.0)
            save_config(cfg)
            dlg.destroy()
            messagebox.showinfo("提示", "价格系数已保存")

        tb.Button(dlg, text="保存", command=save).pack(pady=6)

    def _open_prompt_manager(self):
        """Prompt profile list manager."""
        dlg = tk.Toplevel(self.root)
        dlg.title("提示词管理")
        dlg.geometry("400x350")
        dlg.transient(self.root)
        prompts = load_prompts()

        tree = tb.Treeview(dlg, columns=("name",), show="headings", height=10)
        tree.heading("name", text="提示词档")
        tree.column("name", width=300)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        for name in prompts:
            tree.insert("", tk.END, values=(name,))

        btn_frame = tb.Frame(dlg)
        btn_frame.pack(pady=6)

        def _add():
            n = simpledialog.askstring("新增", "输入提示词档名称:", parent=dlg)
            if n and n not in prompts:
                prompts[n] = ""
                save_prompts(prompts)
                tree.insert("", tk.END, values=(n,))
                self._refresh_prompt_profiles()

        def _edit():
            sel = tree.selection()
            if not sel:
                return
            name = tree.item(sel[0])["values"][0]
            ed_dlg = tk.Toplevel(dlg)
            ed_dlg.title(f"编辑 - {name}")
            ed_dlg.geometry("500x300")
            ed_dlg.transient(dlg)
            t = tk.Text(ed_dlg, wrap=tk.WORD)
            t.insert("1.0", prompts.get(name, ""))
            t.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            tb.Button(ed_dlg, text="保存并关闭",
                       command=lambda: [_save_edit(), ed_dlg.destroy()]).pack(pady=6)

            def _save_edit():
                prompts[name] = t.get("1.0", tk.END).strip()
                save_prompts(prompts)
                self._refresh_prompt_profiles()

        def _delete():
            sel = tree.selection()
            if sel:
                name = tree.item(sel[0])["values"][0]
                if name == "generic":
                    messagebox.showinfo("提示", "generic 是默认档，不可删除")
                    return
                if messagebox.askyesno("确认", f"删除 [{name}]?"):
                    del prompts[name]
                    save_prompts(prompts)
                    tree.delete(sel[0])
                    self._refresh_prompt_profiles()

        def _test_from_manager():
            sel = tree.selection()
            if sel:
                name = tree.item(sel[0])["values"][0]
                self._notebook.select(1)  # Switch to test tab
                self._test_type.set("image")
                self._toggle_test_type()
                self._img_category_var.set(name)
                self._img_prompt.delete("1.0", tk.END)
                self._img_prompt.insert("1.0", prompts.get(name, ""))
                dlg.destroy()

        tb.Button(btn_frame, text="新增", command=_add).pack(side=tk.LEFT, padx=4)
        tb.Button(btn_frame, text="编辑", command=_edit).pack(side=tk.LEFT, padx=4)
        tb.Button(btn_frame, text="删除", command=_delete).pack(side=tk.LEFT, padx=4)
        tb.Button(btn_frame, text="在测试面板打开", command=_test_from_manager).pack(side=tk.LEFT, padx=4)

    def _open_batch_settings(self):
        """Batch size & quantity dialog."""
        cfg = load_config()
        current_batch = cfg.get("batch_size", 10)
        current_qty = cfg.get("default_quantity", 50)
        dlg = tk.Toplevel(self.root)
        dlg.title("处理设置")
        dlg.geometry("280x200")
        dlg.transient(self.root)
        tb.Label(dlg, text="每批处理产品数:").pack(pady=(8, 0))
        batch_var = tk.IntVar(value=current_batch)
        cb = tb.Combobox(dlg, textvariable=batch_var, values=[5, 10, 15, 20, 30], state="readonly")
        cb.pack(pady=2)
        tb.Label(dlg, text="默认库存数量:").pack(pady=(8, 0))
        qty_var = tk.IntVar(value=current_qty)
        tb.Spinbox(dlg, from_=1, to=9999, textvariable=qty_var, width=10).pack(pady=2)

        def save():
            cfg["batch_size"] = batch_var.get()
            cfg["default_quantity"] = qty_var.get()
            save_config(cfg)
            dlg.destroy()

        tb.Button(dlg, text="保存", command=save).pack(pady=10)

    # ============================================================
    # WAVE 4 — task-09: Prompt Manager (integrated above in _open_prompt_manager)
    # ============================================================

    # ============================================================
    # Help
    # ============================================================

    def _open_help(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("使用说明 — 耀我科技上传器 v2.0")
        dlg.geometry("600x500")
        dlg.transient(self.root)
        t = tk.Text(dlg, wrap=tk.WORD, font=("Microsoft YaHei", 10))
        t.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        help_text = """# 耀我科技上传器 v2.0 — 使用说明

## 第一步：配置 API
点击菜单 设置 → API 配置。必填项：DeepSeek Key。
haomingai Key 和 Cloudinary 选填（不填则跳过对应功能）。
代理地址留空则自动检测系统代理。

## 第二步：选择文件
工具栏点击采集表（采集器输出的 Excel）和模板（Gmarket 上架表模板）。
选择提示词档（默认为 generic）。

## 第三步：开始处理
点击"开始" → 选择运行模式 → 程序自动处理。
10 个产品为一批：生成标题/类目/价格 → AI 生图 → 写入 Excel。

## 运行模式
- 全自动运行：无人值守，失败产品标红跳过
- 错误确认模式：每批有失败时暂停，问你继续还是重试

## 提示词测试
右侧切到"测试"标签 → 标题测试：编辑提示词→测试→查看结果→保存。
生图测试：填类目名→编辑提示词→选择参考图产品→测试→保存。
保存后提示词档自动出现在工具栏下拉框中。

## 断点续传
处理中断后，选择同一采集表重新开始，程序自动检测已完成的
ParentSKU 并跳过，只处理剩余的。

## 价格系数
设置 → 价格系数 → 编辑分段和系数 → 全局乘数 → 保存。
公式：中位数 × 分段系数 × 全局乘数，取整到 10 韩元。

## 输出
文件保存在采集表同目录，命名格式：
  <模板名>_<日期>_<序号>.xlsx
序号自动递增，不会覆盖之前的输出。
"""
        t.insert("1.0", help_text.strip())
        t.configure(state="disabled")

    # ============================================================
    # File selection and actions
    # ============================================================

    def _select_source(self):
        path = filedialog.askopenfilename(title="选择采集表", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.source_path = path
            self._src_btn.configure(text=os.path.basename(path))
            self._load_products()

    def _import_more_skus(self):
        """Append products from another Excel to existing list (dedup by ParentSKU)."""
        path = filedialog.askopenfilename(title="选择要导入的采集表", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(min_row=2, values_only=True)
            existing = {p.parent_sku for p in self.products}
            from models import Product
            new_groups = {}
            for row in rows:
                if not row[0]:
                    continue
                psku = str(row[0])
                if psku in existing:
                    continue  # skip duplicates
                if psku not in new_groups:
                    new_groups[psku] = Product(
                        parent_sku=psku,
                        title=str(row[2] or ""),
                        tag=str(row[4] or ""),
                        price=str(row[12] or ""),
                        url=str(row[11] or ""),
                        main_img=str(row[17] or ""),
                    )
                color = str(row[9] or "")
                size = str(row[10] or "")
                price = str(row[12] or "")
                if color and color != "Default" and color not in new_groups[psku].colors:
                    new_groups[psku].colors.append(color)
                if size and size != "One Size" and size not in new_groups[psku].sizes:
                    new_groups[psku].sizes.append(size)
                new_groups[psku].color_sizes.append((color, size, price))
                vi = str(row[40] or "")
                if vi and vi not in new_groups[psku].variant_imgs:
                    new_groups[psku].variant_imgs.append(vi)
                for i in range(17, 38):
                    u = str(row[i] or "")
                    if u and u.startswith("http") and u not in new_groups[psku].extra_imgs:
                        new_groups[psku].extra_imgs.append(u)
            new_list = list(new_groups.values())
            if new_list:
                self.products.extend(new_list)
                self._refresh_list()
                messagebox.showinfo("提示", f"已导入 {len(new_list)} 个新 SKU（跳过 {len(existing)} 个已存在的）")
            else:
                messagebox.showinfo("提示", "没有新的 SKU 可导入（全部已存在）")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def _select_template(self):
        path = filedialog.askopenfilename(title="选择上架表模板", filetypes=[("Excel", "*.xlsx")])
        if path:
            self.template_path = path
            self._tpl_btn.configure(text=os.path.basename(path))

    def _load_products(self):
        """Load product list from source Excel."""
        if not self.source_path:
            return
        try:
            platform = self._platform_var.get()
            from openpyxl import load_workbook
            wb = load_workbook(self.source_path, read_only=True)
            ws = wb.active
            rows = ws.iter_rows(min_row=2, values_only=True)
            from models import Product
            groups = {}
            for row in rows:
                if not row[0]:
                    continue
                psku = str(row[0])
                if psku not in groups:
                    groups[psku] = Product(
                        parent_sku=psku,
                        title=str(row[2] or ""),
                        tag=str(row[4] or ""),
                        price=str(row[12] or ""),
                        url=str(row[11] or ""),
                        main_img=str(row[17] or ""),
                        platform=platform,
                    )
                color = str(row[9] or "")
                size = str(row[10] or "")
                price = str(row[12] or "")
                if color and color != "Default" and color not in groups[psku].colors:
                    groups[psku].colors.append(color)
                if size and size != "One Size" and size not in groups[psku].sizes:
                    groups[psku].sizes.append(size)
                groups[psku].color_sizes.append((color, size, price))
                vi = str(row[40] or "")
                if vi and vi not in groups[psku].variant_imgs:
                    groups[psku].variant_imgs.append(vi)
                for i in range(17, 38):
                    u = str(row[i] or "")
                    if u and u.startswith("http") and u not in groups[psku].extra_imgs:
                        groups[psku].extra_imgs.append(u)
                # AliExpress: read description images from cols 67-86
                if platform == "aliexpress":
                    for i in range(66, 86):
                        u = str(row[i] or "") if i < len(row) else ""
                        if u and u.startswith("http") and u not in groups[psku].desc_images:
                            groups[psku].desc_images.append(u)
            # Dedup (color,size) combos
            for p in groups.values():
                seen, deduped = set(), []
                for cs in p.color_sizes:
                    key = (cs[0], cs[1])
                    if key not in seen:
                        seen.add(key)
                        deduped.append(cs)
                p.color_sizes = deduped
            self.products = list(groups.values())
            # Store original main_img for platform switching
            self._orig_main_img = {p.parent_sku: p.main_img for p in self.products}
            # Apply initial platform setting
            self._apply_platform_images()
            self._refresh_list()
            messagebox.showinfo("提示", f"已加载 {len(self.products)} 个产品")
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def _apply_platform_images(self):
        """Swap main_img based on platform: AliExpress uses 1st variant image."""
        platform = self._platform_var.get()
        for p in self.products:
            p.platform = platform
            orig = self._orig_main_img.get(p.parent_sku, "")
            if platform == "aliexpress" and p.variant_imgs:
                p.main_img = p.variant_imgs[0]
            else:
                p.main_img = orig

    def _on_platform_changed(self, event=None):
        if not self.products:
            return
        self._apply_platform_images()
        self._refresh_list()

    def _refresh_list(self):
        """Rebuild Treeview from self.products."""
        for item in self._tree.get_children():
            self._tree.delete(item)
        for i, prod in enumerate(self.products):
            self._insert_product_row(i, prod)
        self._prog["maximum"] = len(self.products)
        self._prog_label.configure(text=f"0/{len(self.products)}")

    def _insert_product_row(self, i, prod):
        cat_zh = prod.result.get("K_zh", "")
        checked = "☑" if str(i) in self._checked_products else "☐"
        self._tree.insert("", tk.END, iid=str(i),
                          values=(checked, i + 1, prod.parent_sku, prod.status.value,
                                  cat_zh or "",
                                  (prod.ai_title or prod.title)[:30]),
                          tags=(prod.status.name,))

    def update_product_status(self, index, prod):
        """Update a single product's row in the Treeview."""
        if str(index) in self._tree.get_children():
            cat_zh = prod.result.get("K_zh", "")
            checked = "☑" if str(index) in self._checked_products else "☐"
            self._tree.item(str(index),
                            values=(checked, index + 1, prod.parent_sku, prod.status.value,
                                    cat_zh or "",
                                    (prod.ai_title or prod.title)[:30]),
                            tags=(prod.status.name,))

    def _on_tree_click(self, event):
        """Toggle checkbox when clicking the check column."""
        region = self._tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self._tree.identify_column(event.x)
        if col != "#1":  # check column
            return
        item = self._tree.identify_row(event.y)
        if not item:
            return
        if item in self._checked_products:
            self._checked_products.discard(item)
        else:
            self._checked_products.add(item)
        idx = int(item)
        if 0 <= idx < len(self.products):
            self.update_product_status(idx, self.products[idx])

    def _on_tree_right_click(self, event):
        """Right-click context menu: copy SKU, delete product."""
        item = self._tree.identify_row(event.y)
        if not item:
            return
        idx = int(item)
        if not (0 <= idx < len(self.products)):
            return
        prod = self.products[idx]
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="复制 ParentSKU", command=lambda: self.root.clipboard_append(prod.parent_sku))
        menu.add_command(label="复制 AI标题", command=lambda: self.root.clipboard_append(prod.ai_title or prod.title))
        menu.add_separator()
        menu.add_command(label="删除此产品", command=lambda i=idx: self._delete_product(i))
        menu.post(event.x_root, event.y_root)

    def _delete_product(self, idx):
        """Remove a product from the list."""
        if messagebox.askyesno("确认", f"确定删除 #{idx+1} {self.products[idx].parent_sku}？"):
            del self.products[idx]
            self._checked_products.discard(str(idx))
            # Re-index checked products after deletion
            new_checked = set()
            for c in self._checked_products:
                ci = int(c)
                if ci > idx:
                    new_checked.add(str(ci - 1))
                elif ci < idx:
                    new_checked.add(c)
            self._checked_products = new_checked
            self._refresh_list()

    def _regenerate_checked(self):
        """Re-generate main images for all checked products."""
        if not self._checked_products:
            messagebox.showinfo("提示", "请先在列表中勾选要重新生成的产品")
            return
        if self._processing:
            messagebox.showinfo("提示", "处理进行中，请等完成后再重新生成")
            return
        count = len(self._checked_products)
        if not messagebox.askyesno("确认", f"将为 {count} 个勾选产品重新生成主图，是否继续？"):
            return
        self._regen_status.configure(text=f"批量生成 {count} 个...", foreground="#f39c12")
        self._regen_btn.configure(state="disabled")
        threading.Thread(target=self._do_batch_regen, args=(list(self._checked_products),), daemon=True).start()

    def _do_batch_regen(self, items):
        from processor import phase1_title, phase1_category, phase1_price, phase1_w_type, phase1_x_attr, phase1_y_list
        from processor import _gen_main_image, _gen_detail_html, _collect_variant_imgs, _get_category_zh
        from config_manager import load_config, load_prompts, load_categories, load_banned_words
        from api_client import create_storage_provider

        cfg = load_config()
        prompts = load_prompts()
        prompt_key = self._img_profile_var.get()
        prompt_text = prompts.get(prompt_key, prompts.get("generic", ""))
        storage = create_storage_provider(cfg)
        categories = load_categories(os.path.join(SKILL_DIR, "uploader", "카테고리목록(类别代码K列).xls"))
        banned_path = os.path.join(SKILL_DIR, "uploader", "违禁词gmk.txt")
        banned_words = load_banned_words(banned_path) if os.path.exists(banned_path) else []

        prods_to_regen = []
        for item in items:
            idx = int(item)
            if 0 <= idx < len(self.products):
                prods_to_regen.append((idx, self.products[idx]))

        errors = []
        lock = threading.Lock()

        def worker(idx, prod):
            try:
                prod.status = ProductStatus.PHASE1_TITLE
                self.root.after(0, lambda i=idx, p=prod: self.update_product_status(i, p))
                prod.ai_title = phase1_title(prod, banned_words, prompts)

                prod.status = ProductStatus.PHASE1_CATEGORY
                self.root.after(0, lambda i=idx, p=prod: self.update_product_status(i, p))
                cat_path, k, l, m = phase1_category(prod, categories, prompt_key)
                prod.result["K"] = k
                prod.result["L"] = l
                prod.result["M"] = m
                prod.result["K_path"] = cat_path
                prod.result["K_zh"] = _get_category_zh(cat_path)

                prod.status = ProductStatus.PHASE1_PRICE
                prod.result["O"] = prod.result.get("O") or phase1_price(prod, self.products, cfg)
                prod.result["P"] = prod.result["O"]

                prod.status = ProductStatus.PHASE1_DONE
                prod.result["W"] = phase1_w_type(prod)
                prod.result["X"] = phase1_x_attr(prod, prod.result["W"])
                prod.result["Y"] = phase1_y_list(prod, prod.result["W"], prod.result["X"])
                self.root.after(0, lambda i=idx, p=prod: self.update_product_status(i, p))

                prod.status = ProductStatus.PHASE2_MAIN_IMG
                self.root.after(0, lambda i=idx, p=prod: self.update_product_status(i, p))
                prod.result["Z"] = _gen_main_image(prod, prompt_text, storage)
                self.root.after(0, lambda i=idx, p=prod: self.update_product_status(i, p))

                prod.status = ProductStatus.PHASE2_DETAIL
                prod.result["AB"] = _gen_detail_html(prod, self.products, storage)

                prod.status = ProductStatus.PHASE2_VARIANT
                prod.result["AA"] = _collect_variant_imgs(prod)

                prod.status = ProductStatus.DONE
                prod.logs.append(f"{time.strftime('%H:%M:%S')} [重处理] 全部完成")
                self.root.after(0, lambda i=idx, p=prod: (self.update_product_status(i, p), self._show_preview(p)))
            except Exception as e:
                with lock:
                    errors.append((idx, str(e)))
                prod.status = ProductStatus.FAILED
                prod.logs.append(f"{time.strftime('%H:%M:%S')} [重处理] 失败: {e}")

        threads = [threading.Thread(target=worker, args=(idx, prod)) for idx, prod in prods_to_regen]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        total = len(prods_to_regen)
        failed = len(errors)
        done = total - failed
        self.root.after(0, lambda: (self._regen_btn.configure(state="normal"),
                                     self._regen_status.configure(
                                         text=f"全部完成 {total}个" if not failed else f"完成 {done}/{total}, 失败 {failed}",
                                         foreground="#27ae60" if not failed else "#f39c12")))

    def _do_export(self):
        """Export current product list to Excel — always uses latest self.products data."""
        if not self.products:
            messagebox.showinfo("提示", "暂无产品可导出，请先加载采集表")
            return
        if not self.template_path:
            messagebox.showinfo("提示", "请先选择上架表模板")
            return

        path = filedialog.asksaveasfilename(
            title="导出上架表", defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile=f"上架表_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path:
            return

        try:
            from openpyxl import load_workbook
            import shutil
            # Copy template to preserve formatting
            shutil.copy2(self.template_path, path)
            twb = load_workbook(path)
            tws = twb["NEW 일반상품"]
            from processor import write_product_row, init_output_workbook
            _, _, fixed = init_output_workbook(self.template_path, path)
            for i, prod in enumerate(self.products):
                if prod.status == ProductStatus.DONE:
                    write_product_row(tws, i, prod, fixed, 8)
            twb.save(path)
            messagebox.showinfo("提示", f"已导出 {len([p for p in self.products if p.status == ProductStatus.DONE])} 个产品到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".uploader_state.pkl")

    def _save_state(self):
        """Save current state to pickle file for resume."""
        import pickle
        state = {
            "source_path": self.source_path,
            "template_path": self.template_path,
            "output_path": getattr(getattr(self, '_pipeline', None), 'output_path', None),
            "products": self.products,
            "img_profile": self._img_profile_var.get(),
            "title_profile": self._title_profile_var.get(),
            "mode": self._mode_var.get(),
            "stats": self._stats,
        }
        with open(self.STATE_FILE, "wb") as f:
            pickle.dump(state, f)

    @classmethod
    def has_saved_state(cls):
        return os.path.exists(cls.STATE_FILE)

    @classmethod
    def load_saved_state(cls):
        """Return (products, source_path, template_path, output_path, img_profile, title_profile, mode, stats) or None."""
        import pickle
        if not os.path.exists(cls.STATE_FILE):
            return None
        try:
            with open(cls.STATE_FILE, "rb") as f:
                state = pickle.load(f)
            return state
        except Exception:
            return None

    @classmethod
    def clear_saved_state(cls):
        if os.path.exists(cls.STATE_FILE):
            os.remove(cls.STATE_FILE)

    def _on_close(self):
        if self._processing:
            if messagebox.askyesno("确认", "处理进行中，确定退出？"):
                self._stopped = True
                self._pause_event.set()
        # Save state if there are products
        if self.products:
            self._save_state()
        else:
            self.clear_saved_state()
        self.root.destroy()

    # ============================================================
    # Process control stubs (fully built in Wave 6)
    # ============================================================

    def _start_processing(self):
        if not self.source_path:
            messagebox.showinfo("提示", "请先选择采集表")
            return
        if not self.template_path:
            messagebox.showinfo("提示", "请先选择上架表模板")
            return
        if not self.products:
            messagebox.showinfo("提示", "采集表无产品，请重新选择")
            return
        mode = messagebox.askyesnocancel("运行模式",
                                         "选择运行模式:\n\n是 = 全自动运行\n否 = 错误确认模式\n取消 = 返回")
        if mode is None:
            return
        self._mode_var.set("全自动运行" if mode else "错误确认模式")
        self._processing = True
        self._stopped = False
        self._paused = False
        self._pause_event.set()
        self._pause_btn.configure(text="暂 停")

        # Build output path
        date_str = time.strftime("%Y%m%d")
        tpl_stem = os.path.splitext(os.path.basename(self.template_path))[0]
        cfg = load_config()
        base_dir = cfg.get("save_dir") or os.path.dirname(self.source_path)
        counter = 1
        import glob
        existing = glob.glob(os.path.join(base_dir, f"{tpl_stem}_{date_str}_*.xlsx"))
        if existing:
            nums = []
            for f in existing:
                try:
                    nums.append(int(f.rsplit("_", 1)[-1].replace(".xlsx", "")))
                except ValueError:
                    pass
            if nums:
                counter = max(nums) + 1
        output_path = os.path.join(base_dir, f"{tpl_stem}_{date_str}_{counter}.xlsx")

        prompt_key = self._img_profile_var.get()
        from processor import ProcessingPipeline

        def progress_cb(n, total, detail):
            self.root.after(0, lambda n=n, t=total, d=detail: self._update_progress(n, t, d))

        def stat_cb(ds, hm, up):
            self._stats["deepseek"] = ds
            self._stats["haomingai"] = hm
            self._stats["upload"] = up
            self.root.after(0, lambda: self._stat_label.configure(
                text=f"DeepSeek: {ds} | 生图: {hm} | 上传: {up}"))

        # Error callback (runs in worker thread — use wait_variable for sync)
        def error_cb(msg):
            result = [True]  # default continue
            event = threading.Event()
            def _ask():
                ok = messagebox.askyesno("批次错误", msg + "\n\n是否继续下一批？\n是=继续  否=停止")
                result[0] = ok
                event.set()
            self.root.after(0, _ask)
            event.wait()  # block worker until user responds
            return result[0]

        self._pipeline = ProcessingPipeline(
            self.source_path, self.template_path, output_path,
            prompt_key, mode, progress_cb, stat_cb, error_cb)

        # Run in background thread
        self._mode_var.set("处理中...")
        def _run():
            try:
                ds, hm, up = self._pipeline.run(self.products)
                self.root.after(0, lambda: self._on_processing_done(ds, hm, up))
            except Exception as e:
                err = str(e)
                import traceback
                tb = traceback.format_exc()
                self.root.after(0, lambda msg=err, tb=tb: self._on_processing_error(f"{msg}\n\n{tb}"))

        threading.Thread(target=_run, daemon=True).start()

    def _update_progress(self, n, total, detail):
        self._prog["value"] = n
        self._prog_label.configure(text=f"{n}/{total}")
        # Update product status in list
        if 0 <= n - 1 < len(self.products):
            self.update_product_status(n - 1, self.products[n - 1])

    def _on_processing_done(self, ds, hm, up):
        self._processing = False
        self._mode_var.set("处理完成")
        self._pause_btn.configure(text="暂 停", state="disabled")
        messagebox.showinfo("完成", f"处理完成！\n\n"
                            f"DeepSeek调用: {ds}次\n生图: {hm}张\n上传: {up}次\n\n"
                            f"输出: {self._pipeline.output_path}")

    def _on_processing_error(self, err):
        self._processing = False
        self._mode_var.set("处理出错")
        messagebox.showerror("处理出错", err)

    def _toggle_pause(self):
        if not self._processing:
            return
        if self._paused:
            self._paused = False
            self._pause_event.set()
            if hasattr(self, '_pipeline'):
                self._pipeline.resume()
            self._pause_btn.configure(text="暂 停")
        else:
            self._paused = True
            self._pause_event.clear()
            if hasattr(self, '_pipeline'):
                self._pipeline.pause()
            self._pause_btn.configure(text="继 续")

    def _stop(self):
        if self._processing:
            self._stopped = True
            self._pause_event.set()
            if hasattr(self, '_pipeline'):
                self._pipeline.stop()
            self._mode_var.set("已停止")
            self._pause_btn.configure(state="disabled")

    def _filter_failed(self):
        """Show only failed products in the list."""
        for item in self._tree.get_children():
            self._tree.delete(item)
        idx_map = []  # display position → real index
        for i, prod in enumerate(self.products):
            if prod.status == ProductStatus.FAILED:
                idx_map.append(i)
                cat_zh = prod.result.get("K_zh", "")
                checked = "☑" if str(i) in self._checked_products else "☐"
                self._tree.insert("", tk.END, iid=str(i),
                                  values=(checked, i + 1, prod.parent_sku, prod.status.value,
                                          cat_zh or "", (prod.ai_title or prod.title)[:30]),
                                  tags=(prod.status.name,))
        self._mode_var.set(f"筛选失败: {len(idx_map)} 个")

    def _uncheck_all(self):
        """Clear all checkmarks."""
        self._checked_products.clear()
        self._refresh_list()
        messagebox.showinfo("提示", "已取消全部勾选")

    def _check_all_failed(self):
        """Check all failed products in the selection/refresh."""
        failed_count = 0
        for i, prod in enumerate(self.products):
            if prod.status == ProductStatus.FAILED:
                self._checked_products.add(str(i))
                failed_count += 1
        self._refresh_list()
        messagebox.showinfo("提示", f"已勾选 {failed_count} 个失败产品，可点击「重新生成选中」")

    def _clear_list(self):
        if not self.products:
            return
        if self._processing:
            messagebox.showinfo("提示", "处理进行中，请先停止再清空列表")
            return
        if messagebox.askyesno("确认", f"确定清空当前列表？\n\n{len(self.products)} 个产品将被移除。"):
            self._checked_products.clear()
            self.products = []
            self.source_path = None
            self._src_btn.configure(text="无")
            self._refresh_list()
            self._mode_var.set("等待开始...")
            self._prog["value"] = 0
            self._prog_label.configure(text="0/0")
            self._stat_label.configure(text="DeepSeek: 0 | 生图: 0 | 上传: 0")
            self._stats = {"deepseek": 0, "haomingai": 0, "upload": 0}

    def _retry_all_failed(self):
        if hasattr(self, '_pipeline') and self._pipeline.errors:
            self._start_processing()
        else:
            messagebox.showinfo("提示", "没有失败的产品需要重试")

    def run(self):
        """Launch the Tkinter main loop."""
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.mainloop()

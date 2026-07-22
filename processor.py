# Author: Administrator
# Created: 2026-05-25
"""Processing pipeline for 耀我科技上传器 v2.0 — batch scheduler, Phase 1/2, Excel write."""

import os
import re
import json
import time
import threading
import io
from dataclasses import dataclass
from urllib.parse import urlsplit, urlunsplit
from openpyxl import load_workbook

from models import Product, Batch, ProductStatus
from config_manager import load_config, load_prompts, load_categories, load_banned_words
from config_manager import load_category_zh, save_category_zh
from api_client import deepseek_chat, generate_image, download_image
from api_client import create_storage_provider
from image_utils import ensure_marketplace_image_spec
from store_profiles import (
    category_fields,
    load_store_profiles,
    match_store_category,
    resolve_fields,
)

SKILL_DIR = os.path.dirname(os.path.abspath(__file__))

# ---- Helpers ----

def _color_to_korean(color, translate_fn):
    """Translate color name to Korean if needed."""
    if not color:
        return color
    if re.match(r'^[가-힣]+$', color):
        return color
    try:
        return translate_fn(color, target="ko")
    except:
        return color


def _strip_particles(word):
    """Strip Korean particles from a keyword."""
    particles = ['으로', '로', '의', '은', '는', '이', '가', '을', '를',
                 '에', '에게', '도', '만', '와', '과', '하고', '이나', '나']
    for p in particles:
        if word.endswith(p) and len(word) > len(p) + 1:
            return word[:-len(p)]
    return word


def _url_stem(url):
    """Strip _thumbnail suffix from URL for dedup."""
    if not url:
        return url
    return re.sub(r'_thumbnail[^/]*(\.[\w]+)$', r'\1', url)


# ---- Excel Read/Write ----

def load_products(source_path, platform="shein"):
    """Load products from scraped Excel, grouping by ParentSKU.

    For AliExpress: also reads description images from cols 67-86.
    """
    wb = load_workbook(source_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)  # generator, not list
    groups = {}
    for row in rows:
        if not row[0]:
            continue
        psku = str(row[0])
        if psku not in groups:
            groups[psku] = Product(
                parent_sku=psku,
                title=str(row[2] or ""),
                tag=str(row[4] or ""),
                brand=str(row[5] or ""),
                price=str(row[12] or ""),
                url=str(row[11] or ""),
                main_img=str(row[17] or ""),
                source_main_img=str(row[17] or ""),
                platform=platform,
            )
        color = str(row[9] or "")
        size = str(row[10] or "")
        price = str(row[12] or "")
        if color and color != "Default" and color not in groups[psku].colors:
            groups[psku].colors.append(color)
        if size and size != "One Size" and size not in groups[psku].sizes:
            groups[psku].sizes.append(size)
        groups[psku].color_sizes.append((color, size, price))
        vi = str(row[40] or "")
        if vi and vi not in groups[psku].variant_imgs:
            groups[psku].variant_imgs.append(vi)
        for i in range(17, 38):
            u = str(row[i] or "")
            if u and u.startswith("http") and u not in groups[psku].extra_imgs:
                groups[psku].extra_imgs.append(u)
        # AliExpress: read description images from cols 67-86 (0-indexed: 66-85)
        if platform == "aliexpress":
            for i in range(66, 86):
                u = str(row[i] or "") if i < len(row) else ""
                if u and u.startswith("http") and u not in groups[psku].desc_images:
                    groups[psku].desc_images.append(u)
    # Dedup (color,size) combos within each product
    for p in groups.values():
        seen, deduped = set(), []
        for cs in p.color_sizes:
            key = (cs[0], cs[1])
            if key not in seen:
                seen.add(key)
                deduped.append(cs)
        p.color_sizes = deduped
    # AliExpress: main_img = 1st variant image
    if platform == "aliexpress":
        for p in groups.values():
            if p.variant_imgs:
                p.main_img = p.variant_imgs[0]
    return list(groups.values())


def detect_completed(output_path):
    """Return set of ParentSKUs already written to output Excel."""
    if not output_path or not os.path.exists(output_path):
        return set()
    try:
        wb = load_workbook(output_path, read_only=True)
        ws = wb["NEW 일반상품"]
        done = set()
        for row in ws.iter_rows(min_row=8, values_only=True):
            # Column B is ParentSKU-like, but we check column E (AI title) to confirm it was processed
            title_cell = row[4] if len(row) > 4 else None
            sku_cell = row[0] if row[0] else None
            if title_cell and str(title_cell).strip():
                # Use row index as indicator; store SKU from col E or col B
                done.add(str(row[1]) if row[1] else "")
        return done
    except:
        return set()


def init_output_workbook(template_path, output_path, cfg=None, store_id="", category_id=""):
    """Copy template to output. Return (workbook, sheet, start_row, fixed_dict)."""
    if cfg is None:
        cfg = load_config()
    twb = load_workbook(template_path)
    tws = twb["NEW 일반상품"]
    tpl_row = 8
    from openpyxl.utils import get_column_letter

    template_fields = {
        get_column_letter(column): tws.cell(row=tpl_row, column=column).value
        for column in range(1, 67)
    }
    fixed = resolve_fields(template_fields, store_id, category_id)
    quantity = cfg.get("default_quantity")
    if quantity not in (None, ""):
        fixed["U"] = quantity
        fixed["V"] = quantity
    return twb, tws, fixed


def write_product_row(tws, row_idx, prod, fixed, tpl_start_row, category_fixed=None):
    """Write one product row into the template sheet."""
    from openpyxl.utils import column_index_from_string

    r = tpl_start_row + row_idx
    merged_fixed = dict(fixed or {})
    merged_fixed.update(category_fixed or {})
    for column, value in merged_fixed.items():
        tws.cell(row=r, column=column_index_from_string(column), value=value)
    # A: seq
    tws.cell(row=r, column=1, value=r - 5)
    # E: AI title
    tws.cell(row=r, column=5, value=prod.ai_title)
    # K/L/M: category
    tws.cell(row=r, column=11, value=prod.result.get("K", ""))
    tws.cell(row=r, column=12, value=prod.result.get("L", ""))
    tws.cell(row=r, column=13, value=prod.result.get("M", ""))
    # O/P: price
    p = prod.result.get("O", prod.price)
    tws.cell(row=r, column=15, value=p)
    tws.cell(row=r, column=16, value=p)
    # W/X/Y: type/attr/color_size
    tws.cell(row=r, column=23, value=prod.result.get("W", "미사용"))
    tws.cell(row=r, column=24, value=prod.result.get("X", "색상"))
    tws.cell(row=r, column=25, value=prod.result.get("Y", ""))
    # Z: main image URL
    tws.cell(row=r, column=26, value=prod.result.get("Z", prod.main_img))
    # AA: variant images
    tws.cell(row=r, column=27, value=_dedupe_csv_urls(prod.result.get("AA", "")))
    # AB: detail HTML
    tws.cell(row=r, column=28, value=prod.result.get("AB", ""))


# ---- Phase 1: Serial per-product ----

_SUBJECT_PROFILE_FIELDS = (
    "sold_object",
    "sold_object_ko",
    "sold_object_zh",
    "buyer_receives",
    "primary_function",
    "category_terms_ko",
    "category_terms_zh",
    "referenced_objects",
    "attributes",
    "evidence",
    "confidence",
)


def _parse_json_object_response(text):
    """Return the first JSON object in a model response."""
    raw = str(text or "")
    decoder = json.JSONDecoder()
    for index, char in enumerate(raw):
        if char != "{":
            continue
        try:
            value, _ = decoder.raw_decode(raw[index:])
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
        if isinstance(value, dict):
            return value
    return {}


def _subject_list(value):
    if isinstance(value, (list, tuple, set)):
        items = value
    elif value:
        items = [value]
    else:
        items = []
    result = []
    for item in items:
        text = str(item or "").strip()
        if text and text not in result:
            result.append(text)
    return result


def _normalize_subject_profile(value, fallback_title=""):
    """Normalize model output without using product-specific dictionaries."""
    raw = value if isinstance(value, dict) else {}
    profile = {}
    for key in ("sold_object", "sold_object_ko", "sold_object_zh",
                "buyer_receives", "primary_function"):
        profile[key] = str(raw.get(key, "") or "").strip()
    for key in ("category_terms_ko", "category_terms_zh", "referenced_objects",
                "attributes", "evidence"):
        profile[key] = _subject_list(raw.get(key))
    try:
        profile["confidence"] = max(0.0, min(1.0, float(raw.get("confidence", 0) or 0)))
    except (TypeError, ValueError):
        profile["confidence"] = 0.0
    if not any(profile.get(key) for key in ("sold_object", "sold_object_ko", "sold_object_zh")):
        profile["sold_object"] = str(fallback_title or "").strip()
        profile["confidence"] = 0.0
    return profile


def _subject_from_model_object(parsed, fallback_title=""):
    subject = parsed.get("subject") if isinstance(parsed, dict) else None
    if not isinstance(subject, dict):
        subject = {
            key: parsed.get(key)
            for key in _SUBJECT_PROFILE_FIELDS
            if isinstance(parsed, dict) and key in parsed
        }
    return _normalize_subject_profile(subject, fallback_title=fallback_title)


def _subject_prompt_rules():
    return (
        "In the same response, identify the actual item being sold. The sold item is "
        "the physical object the buyer receives in the parcel. Distinguish it from "
        "objects that it stores, supports, protects, connects to, fits, controls, or is "
        "shown with. Do not treat materials, origins, dimensions, styles, compatible "
        "objects, or usage scenes as the sold item. Use semantic understanding only; "
        "do not rely on a fixed product keyword list. Evidence must be copied from the "
        "source title. Return Korean and Chinese category-style names when possible. "
        "Also provide several short taxonomy-style synonyms in category_terms_ko and "
        "category_terms_zh. These terms must describe the sold item itself, never a "
        "referenced or compatible object."
    )


def _brand_only_title_and_subject(prod, banned_words):
    title = str(prod.title or "")
    hints = _brand_hint_words(title, banned_words)
    prompt = (
        "Perform two tasks for this ecommerce product in one response.\n"
        "1. Identify only brand, trademark, store, company, or manufacturer words in "
        "the title. Do not classify colors, specifications, materials, quantities, "
        "functions, or product names as brands.\n"
        f"2. {_subject_prompt_rules()}\n\n"
        "Return JSON only with this exact shape:\n"
        '{"brand_words":[],"subject":{"sold_object":"","sold_object_ko":"",'
        '"sold_object_zh":"","buyer_receives":"","primary_function":"",'
        '"category_terms_ko":[],"category_terms_zh":[],'
        '"referenced_objects":[],"attributes":[],"evidence":[],"confidence":0.0}}\n\n'
        f"Source title: {title}\n"
        f"Auxiliary raw attributes (may be noisy): {getattr(prod, 'tag', '')}\n"
        f"Possible brand hints: {', '.join(hints) if hints else 'none'}"
    )
    result = deepseek_chat(prompt, max_tokens=500, temp=0)
    parsed = _parse_json_object_response(result)
    if parsed:
        brand_words = _clean_brand_word_list(parsed.get("brand_words", []), title)
        profile = _subject_from_model_object(parsed, fallback_title=title)
    else:
        # Backward compatibility for old/plain model responses and existing configs.
        brand_words = _parse_brand_words_response(result, title)
        profile = _normalize_subject_profile({}, fallback_title=title)
    prod.subject_profile = profile
    return _limit_title_length(remove_brand_words_from_title(title, brand_words))

def phase1_title(prod, banned_words, prompts, title_mode="ai_rewrite"):
    """Generate or clean title according to the selected title mode."""
    if _is_brand_only_title_mode(title_mode):
        return _brand_only_title_and_subject(prod, banned_words)

    is_multi = len(prod.colors) > 1
    color_note = ""
    if is_multi:
        color_note = "CRITICAL: This product has MULTIPLE colors. Do NOT mention any color in the title."
    else:
        color_note = "You may include the color keyword in the title."

    title_tpl = prompts.get("title", "")
    if "{product_title}" in title_tpl:
        prompt = title_tpl.format(color_note=color_note, product_title=prod.title,
                                  banned_words=", ".join(banned_words))
    else:
        prompt = f"""你是韩国电商标题优化专家。生成Gmarket商品标题：
- CRITICAL: 纯韩文标题必须45字符以内
- 极致本土化用语，韩国消费者使用的自然表达
- 包含高流量关键词
- 禁止：品牌名、年/月/日
- {color_note}
- 只输出标题文本本身，不要任何解释。

产品：{prod.title}
违禁词：{', '.join(banned_words)}"""
    combined_prompt = (
        "Perform the title task below and sold-item analysis in one response. Any "
        "instruction below that asks for plain-text-only output is superseded by the "
        "JSON output format at the end.\n\n"
        f"TITLE TASK:\n{prompt}\n\n"
        f"SOLD-ITEM TASK:\n{_subject_prompt_rules()}\n\n"
        "Return JSON only with this exact shape:\n"
        '{"title":"","subject":{"sold_object":"","sold_object_ko":"",'
        '"sold_object_zh":"","buyer_receives":"","primary_function":"",'
        '"category_terms_ko":[],"category_terms_zh":[],'
        '"referenced_objects":[],"attributes":[],"evidence":[],"confidence":0.0}}\n\n'
        f"Auxiliary raw attributes (may be noisy): {getattr(prod, 'tag', '')}"
    )
    result = deepseek_chat(combined_prompt, max_tokens=500, temp=0.7)
    parsed = _parse_json_object_response(result)
    if parsed and str(parsed.get("title", "") or "").strip():
        title = str(parsed.get("title", "") or "").strip()
        prod.subject_profile = _subject_from_model_object(parsed, fallback_title=prod.title)
    else:
        title = str(result or "").strip()
        prod.subject_profile = _normalize_subject_profile({}, fallback_title=prod.title)
    return _limit_title_length(title)


def phase1_title_brand_only(title, banned_words):
    """Ask DeepSeek which words are brands, then remove only those words locally."""
    brand_words = identify_brand_words_in_title(title, banned_words)
    return _limit_title_length(remove_brand_words_from_title(title, brand_words))


def _limit_title_length(title, max_len=45):
    title = str(title or "")
    if len(title) <= max_len:
        return title

    cut = title[:max_len].rstrip()
    if not cut:
        return cut

    separators = r"\s,.;:!?，。；：！？、/\\|+\-–—_~·()\[\]{}<>《》「」『』"
    if max_len < len(title) and re.match(f"[{separators}]", title[max_len]):
        return re.sub(f"[{separators}]+$", "", cut).rstrip()

    matches = list(re.finditer(f"[{separators}]+", cut))
    if matches:
        last = matches[-1]
        min_keep = max_len - 18
        if last.start() >= min_keep:
            candidate = cut[:last.start()].rstrip()
            if candidate:
                return re.sub(f"[{separators}]+$", "", candidate).rstrip()

    return cut


def identify_brand_words_in_title(title, banned_words):
    title = str(title or "")
    hints = _brand_hint_words(title, banned_words)
    hint_text = ", ".join(hints) if hints else "无"
    prompt = f"""请只识别下面商品标题里的品牌名、商标名、店铺名或公司名。
不要改写标题，不要翻译标题，不要删除颜色、规格、材质、品类、数量、用途词。
只输出 JSON 数组，例如 ["ZARA", "JBL"]；如果没有品牌名，输出 []。

商品标题：{title}
可疑品牌参考词：{hint_text}"""
    result = deepseek_chat(prompt, max_tokens=120, temp=0)
    return _parse_brand_words_response(result, title)


def remove_brand_words_from_title(title, banned_words):
    """Remove configured brand/forbidden words while leaving title text otherwise intact."""
    cleaned = str(title or "")
    for word in _iter_brand_words(banned_words):
        pattern = re.compile(r"(?<![A-Za-z0-9])" + re.escape(word) + r"(?![A-Za-z0-9])", re.IGNORECASE)
        cleaned = pattern.sub(" ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"\s+([,.;:!?，。；：！？])", r"\1", cleaned)
    return cleaned


def _is_brand_only_title_mode(title_mode):
    return str(title_mode or "").strip() in ("brand_only", "仅去品牌")


def _brand_hint_words(title, banned_words):
    title_text = str(title or "")
    hints = []
    for word in _iter_brand_words(banned_words):
        pattern = re.compile(r"(?<![A-Za-z0-9])" + re.escape(word) + r"(?![A-Za-z0-9])", re.IGNORECASE)
        if pattern.search(title_text):
            hints.append(word)
        if len(hints) >= 30:
            break
    return hints


def _parse_brand_words_response(text, title):
    raw = str(text or "").strip()
    if not raw:
        return []
    json_match = re.search(r"\[[\s\S]*\]", raw)
    if json_match:
        try:
            data = json.loads(json_match.group(0))
            if isinstance(data, list):
                return _clean_brand_word_list(data, title)
        except Exception:
            pass
    if raw.strip().lower() in ("[]", "none", "no brand", "no brands", "n/a"):
        return []
    if raw.strip() in ("无", "没有", "无品牌", "没有品牌"):
        return []
    raw = re.sub(r"^[A-Za-z\u4e00-\u9fff\s]*(?:品牌|brand|brands)\s*[:：]\s*", "", raw, flags=re.IGNORECASE)
    parts = re.split(r"[,，;；\n\r]+", raw)
    return _clean_brand_word_list(parts, title)


def _clean_brand_word_list(words, title):
    title_text = str(title or "").strip()
    cleaned = []
    seen = set()
    for item in words or []:
        word = str(item).strip().strip("\"'`[]()（）")
        word = re.sub(r"^\s*[-*•]\s*", "", word).strip()
        if not word or word in ("无", "没有", "[]"):
            continue
        if word == title_text:
            continue
        key = word.casefold()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(word)
    return cleaned


def _iter_brand_words(banned_words):
    seen = set()
    for item in banned_words or []:
        for part in re.split(r"[,，\n\r]+", str(item)):
            word = part.strip()
            if not word or len(word) < 2:
                continue
            key = word.casefold()
            if key in seen:
                continue
            seen.add(key)
            yield word


# ---- Profile key to Korean translation (DeepSeek + JSON cache) ----
import os as _os, json as _json

def _load_profile_ko():
    p = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "profile_ko.json")
    if _os.path.exists(p):
        try:
            with open(p, 'r', encoding='utf-8') as f:
                return _json.load(f)
        except:
            pass
    return {}

def _save_profile_ko(cache):
    p = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "profile_ko.json")
    with open(p, 'w', encoding='utf-8') as f:
        _json.dump(cache, f, ensure_ascii=False, indent=2)

def _profile_to_korean(profile_key):
    """Translate a profile key (any language) to Korean via DeepSeek, with JSON cache."""
    cache = _load_profile_ko()
    if profile_key in cache:
        return cache[profile_key]
    try:
        prompt = "将以下电商品类关键词翻译为韩文。只输出一个韩文单词，不要解释，不要标点。\n关键词：" + profile_key + "\n韩文："
        ko = deepseek_chat(prompt, max_tokens=50, temp=0.3).strip()
        if ko and len(ko) > 1:
            cache[profile_key] = ko
            _save_profile_ko(cache)
            return ko
    except:
        pass
    return ""


def _find_category_parent(categories, profile_key):
    """Find the deepest common ancestor of all category paths matching profile_key.

    Translates profile_key to Korean via DeepSeek (cached), then searches
    the category table. Works for any language (Chinese, English, Korean).

    Steps:
      1. Translate profile_key to Korean (cached)
      2. Collect ALL paths where any segment matches the Korean term
      3. Compute the longest common prefix of all matched paths
      4. That common ancestor is the broad category scope
    """
    search_terms = []
    kr = _profile_to_korean(profile_key)
    if kr:
        search_terms.append(kr)
    if profile_key != kr:
        search_terms.append(profile_key)

    matched_paths = []
    for path in categories:
        segments = path.split(">")
        for seg in segments:
            if seg in search_terms:
                matched_paths.append(segments)
                break

    if not matched_paths:
        return set()

    min_len = min(len(p) for p in matched_paths)
    common = []
    for i in range(min_len):
        seg_i = matched_paths[0][i]
        if all(p[i] == seg_i for p in matched_paths):
            common.append(seg_i)
        else:
            break

    if not common:
        return set()

    return {">".join(common)}



_CATEGORY_ZH_MAP = None
_CATEGORY_SEMANTIC_FEATURES = {}


def _category_zh_map():
    global _CATEGORY_ZH_MAP
    if _CATEGORY_ZH_MAP is None:
        try:
            _CATEGORY_ZH_MAP = load_category_zh()
        except Exception:
            _CATEGORY_ZH_MAP = {}
    return _CATEGORY_ZH_MAP


def _compact_semantic_text(value):
    return re.sub(r"[^\w]+", "", str(value or "").casefold(), flags=re.UNICODE)


def _semantic_tokens(value):
    return [
        token for token in re.findall(r"[^\W_]+", str(value or "").casefold(), re.UNICODE)
        if token
    ]


def _character_ngrams(value, size=2):
    text = _compact_semantic_text(value)
    if not text:
        return set()
    if len(text) <= size:
        return {text}
    return {text[index:index + size] for index in range(len(text) - size + 1)}


def _semantic_features(value):
    return (
        _compact_semantic_text(value),
        frozenset(_semantic_tokens(value)),
        frozenset(_character_ngrams(value)),
    )


def _semantic_similarity_features(left, right):
    left_compact, left_tokens, left_grams = left
    right_compact, right_tokens, right_grams = right
    if not left_compact or not right_compact:
        return 0.0
    if left_compact == right_compact:
        return 1.0

    shorter = min(len(left_compact), len(right_compact))
    longer = max(len(left_compact), len(right_compact))
    containment = 0.0
    if left_compact in right_compact or right_compact in left_compact:
        containment = 0.72 + 0.28 * (shorter / longer)

    token_overlap = 0.0
    if left_tokens and right_tokens:
        token_overlap = len(left_tokens & right_tokens) / len(left_tokens | right_tokens)
    gram_overlap = 0.0
    if left_grams and right_grams:
        gram_overlap = len(left_grams & right_grams) / len(left_grams | right_grams)
    return max(containment, token_overlap, gram_overlap)


def _semantic_similarity(left, right):
    """Language-agnostic lexical similarity used only for candidate recall."""
    return _semantic_similarity_features(_semantic_features(left), _semantic_features(right))


def _category_semantic_targets(path, path_zh):
    key = (path, path_zh)
    cached = _CATEGORY_SEMANTIC_FEATURES.get(key)
    if cached is not None:
        return cached
    leaf = path.split(">")[-1]
    leaf_zh = path_zh.split(">")[-1] if path_zh else ""
    leaf_targets = [_semantic_features(leaf)]
    path_targets = [_semantic_features(path)]
    if leaf_zh:
        leaf_targets.append(_semantic_features(leaf_zh))
    if path_zh:
        path_targets.append(_semantic_features(path_zh))
    cached = (tuple(leaf_targets), tuple(path_targets))
    _CATEGORY_SEMANTIC_FEATURES[key] = cached
    return cached


def _profile_identity_texts(profile):
    texts = []
    for key in ("sold_object_ko", "sold_object_zh", "sold_object", "buyer_receives"):
        text = str(profile.get(key, "") or "").strip()
        if text and text not in texts:
            texts.append(text)
    for key in ("category_terms_ko", "category_terms_zh"):
        for text in _subject_list(profile.get(key)):
            if text not in texts:
                texts.append(text)
    return texts


def _subject_profile_is_usable(profile):
    if not isinstance(profile, dict) or not _profile_identity_texts(profile):
        return False
    try:
        return float(profile.get("confidence", 0) or 0) > 0
    except (TypeError, ValueError):
        return False


def _score_categories_from_subject(profile, categories):
    """Recall candidates from the model's semantic-role output, without product rules."""
    identity_texts = _profile_identity_texts(profile)
    identity_features = [_semantic_features(text) for text in identity_texts]
    function_text = str(profile.get("primary_function", "") or "").strip()
    function_features = _semantic_features(function_text)
    referenced = _subject_list(profile.get("referenced_objects"))
    reference_features = [_semantic_features(text) for text in referenced]
    zh_map = _category_zh_map()
    scored = []
    for path, codes in categories.items():
        path_zh = str(zh_map.get(path, "") or "")
        leaf_targets, path_targets = _category_semantic_targets(path, path_zh)

        leaf_match = max(
            (_semantic_similarity_features(query, target)
             for query in identity_features for target in leaf_targets),
            default=0.0,
        )
        path_match = max(
            (_semantic_similarity_features(query, target)
             for query in identity_features for target in path_targets),
            default=0.0,
        )
        function_match = max(
            (_semantic_similarity_features(function_features, target)
             for target in path_targets if function_text),
            default=0.0,
        )
        reference_match = max(
            (_semantic_similarity_features(reference, target)
             for reference in reference_features for target in leaf_targets),
            default=0.0,
        )

        score = leaf_match * 0.72 + path_match * 0.20 + function_match * 0.08
        if (referenced and reference_match >= 0.65
                and reference_match >= leaf_match * 0.75):
            score -= reference_match * 0.65
        if score > 0.12:
            scored.append((score, path, codes))
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored


def _score_categories_legacy(title, tag, categories, profile_key=""):
    """Backward-compatible recall when an old/plain model response has no subject."""
    broad_parents = _find_category_parent(categories, profile_key) if profile_key else set()

    def extract_keys(text):
        keys = []
        for word in re.findall(r'[가-힣\w]+', text):
            stripped = _strip_particles(word)
            if len(stripped) > 1 and stripped not in keys:
                keys.append(stripped)
        return keys

    tag_keys = extract_keys(tag)
    title_keys = extract_keys(title)
    overlap_keys = [key for key in tag_keys if key in title_keys]
    tag_only = [key for key in tag_keys if key not in overlap_keys]
    title_only = [
        key for key in title_keys if key not in overlap_keys and key not in tag_keys
    ]
    scored = []
    for path, codes in categories.items():
        score = 0
        segments = path.split(">")
        last_seg = segments[-1] if segments else ""

        def score_keyword(keyword, exact, partial, upper, upper_partial):
            value = 0
            if keyword == last_seg:
                value += len(keyword) * exact
            elif keyword in last_seg and len(keyword) > 1:
                value += len(keyword) * partial
            else:
                for level, segment in enumerate(segments[:-1]):
                    if keyword == segment:
                        value += len(keyword) * (level + 1) * upper
                    elif keyword in segment and len(keyword) > 1:
                        value += len(keyword) * (level + 1) * upper_partial
            return value

        for prefix in broad_parents:
            if path.startswith(prefix):
                score += 50000
        for keyword in overlap_keys:
            score += score_keyword(keyword, 25000, 12500, 50, 25)
        for keyword in tag_only:
            score += score_keyword(keyword, 20000, 10000, 40, 20)
        for keyword in title_only:
            score += score_keyword(keyword, 15000, 7500, 30, 15)
        if score > 0:
            scored.append((score, path, codes))
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored


def _select_category_with_deepseek(prod, profile, candidates):
    zh_map = _category_zh_map()
    prompt_candidates = []
    by_id = {}
    for index, (_, path, codes) in enumerate(candidates, 1):
        candidate_id = f"C{index:02d}"
        by_id[candidate_id] = (path, codes)
        prompt_candidates.append({
            "candidate_id": candidate_id,
            "path_ko": path,
            "path_zh": zh_map.get(path, ""),
        })
    prompt = (
        "You are selecting a Korean marketplace category. Classify the physical item "
        "the buyer receives, not an object it stores, supports, protects, connects to, "
        "fits, controls, or is shown with. Materials, origins, dimensions, styles, and "
        "usage scenes must not determine the category. Use only the supplied candidates "
        "and copy one candidate_id exactly. Return JSON only with candidate_id, "
        "same_sold_object, confidence, and evidence. same_sold_object must be true only "
        "when the selected category describes the sold item itself.\n\n"
        f"Sold-item analysis: {json.dumps(profile, ensure_ascii=False)}\n"
        f"Candidates: {json.dumps(prompt_candidates, ensure_ascii=False)}"
    )
    prod.result["_category_deepseek_calls"] = (
        int(prod.result.get("_category_deepseek_calls", 0) or 0) + 1
    )
    response = deepseek_chat(prompt, max_tokens=300, temp=0.1)
    parsed = _parse_json_object_response(response)
    if parsed:
        candidate_id = str(parsed.get("candidate_id", "") or "").strip()
        same_object = parsed.get("same_sold_object", True)
        if candidate_id in by_id and same_object is not False:
            return by_id[candidate_id]
        return None
    matched = str(response or "").strip()
    for _, path, codes in candidates:
        if matched == path:
            return path, codes
    return None


def phase1_category(prod, categories, profile_key=""):
    """Match the actual sold object to an uploadable ESM/Gmarket category."""
    prod.result["_category_deepseek_calls"] = 0
    categories = {
        path: codes for path, codes in categories.items()
        if codes.get("esm_code") and codes.get("auction") and codes.get("gmarket")
    }
    if not categories:
        return "", "", "", ""

    profile = getattr(prod, "subject_profile", {}) or {}
    using_subject = _subject_profile_is_usable(profile)
    if using_subject:
        scored = _score_categories_from_subject(profile, categories)
    else:
        scored = _score_categories_legacy(prod.title, prod.tag, categories, profile_key)
        profile = _normalize_subject_profile({}, fallback_title=prod.title)
    if not scored:
        return "", "", "", ""

    try:
        subject_confidence = float(profile.get("confidence", 0) or 0)
    except (TypeError, ValueError):
        subject_confidence = 0.0
    high_confidence_subject_match = (
        using_subject
        and subject_confidence >= 0.85
        and scored[0][0] >= 0.72
    )
    clear_winner = (
        len(scored) == 1
        or scored[0][0] > scored[1][0] * 1.5
        or high_confidence_subject_match
    )
    selected = (scored[0][1], scored[0][2]) if clear_winner else None
    if selected is None:
        candidates = scored[:30 if using_subject else 15]
        try:
            selected = _select_category_with_deepseek(prod, profile, candidates)
        except Exception:
            selected = None
    if selected is None:
        return "", "", "", ""

    matched_path, best = selected
    return (
        matched_path,
        best.get("esm_code", ""),
        str(best.get("auction", "")),
        str(best.get("gmarket", "")),
    )


def ensure_upload_category_codes(esm_code, auction_code, gmarket_code):
    """Raise a clear error instead of exporting a row with unusable A/G codes."""
    if esm_code and auction_code and gmarket_code:
        return
    raise ValueError("未找到同时具备 A/G 平台类目码的可上架类目，请手动选择类目后重试")


def phase1_price(prod, all_products, cfg):
    """Calculate selling price from median of same ParentSKU."""
    psku = prod.parent_sku
    prices = []
    for p in all_products:
        if p.parent_sku == psku:
            for _, _, pr in p.color_sizes:
                try:
                    prices.append(float(pr))
                except ValueError:
                    pass
    if not prices:
        return prod.price or "0"

    prices.sort()
    median = prices[len(prices) // 2]

    coeffs = cfg.get("price_coefficients",
                      [[0, 10000, 1.8], [10000, 20000, 1.8], [20000, 30000, 1.5],
                       [30000, 50000, 1.3], [50000, 999999999, 1.2]])
    multiplier = cfg.get("global_multiplier", 1.0)

    coefficient = 1.8
    for lo, hi, co in coeffs:
        if lo <= median < hi:
            coefficient = co
            break

    final = median * coefficient * multiplier
    final = round(final / 10) * 10
    return str(int(final))


def phase1_w_type(prod):
    n_colors = len(prod.colors)
    n_sizes = len(prod.sizes)
    if n_colors <= 1 and n_sizes <= 1:
        return "미사용"
    elif n_colors > 1 and n_sizes <= 1:
        return "단독형"
    elif n_colors <= 1 and n_sizes > 1:
        return "단독형"
    else:
        return "2개조합형"


def phase1_x_attr(prod, w):
    n_colors = len(prod.colors)
    n_sizes = len(prod.sizes)
    if w == "미사용":
        return "색상"
    elif w == "단독형":
        return "색상" if n_colors > 1 else "사이즈"
    else:
        return "색상,사이즈"


# Cache for DeepSeek abbreviations (shared across all products)
_ABBREV_CACHE = {}

def _abbreviate_if_long(name):
    """If *name* exceeds 50 UTF-8 bytes, ask DeepSeek for a short Korean abbreviation.
    Returns the abbreviated string (or original if short enough / API fails).
    """
    if len(name.encode('utf-8')) <= 50:
        return name
    if name in _ABBREV_CACHE:
        return _ABBREV_CACHE[name]
    try:
        prompt = (
            "다음 옵션명을 50바이트(UTF-8) 이내의 짧은 한국어로 축약해 주세요.\n"
            "원래 의미를 유지하고, 핵심 키워드만 남기세요.\n"
            "설명 없이 축약된 텍스트만 출력하세요.\n"
            f"옵션명: {name}\n축약:"
        )
        short = deepseek_chat(prompt, max_tokens=80, temp=0.2).strip()
        # If API returned something weird, keep original
        if short and 1 < len(short) < len(name):
            _ABBREV_CACHE[name] = short
            return short
    except Exception:
        pass
    _ABBREV_CACHE[name] = name  # Don't retry failures
    return name


def phase1_y_list(prod, w, x):
    from config_manager import load_config
    qty = load_config().get("default_quantity", 50)
    size_noise = ["차트", "사이즈", "표를", "cm", "inch", "길이", "사이즈표", "상세"]
    lines = []
    for color, size, price in prod.color_sizes:
        color_kr = _abbreviate_if_long(color)
        size_clean = size
        for noise in size_noise:
            if noise.lower() in str(size).lower():
                size_clean = ""
                break
        size_clean = _abbreviate_if_long(size_clean) if size_clean else size_clean

        if x == "색상,사이즈":
            lines.append(f"{color_kr},{size_clean},정상,노출,{qty},{qty}")
        elif x == "색상":
            lines.append(f"{color_kr},정상,노출,{qty},{qty}")
        elif x == "사이즈":
            lines.append(f"{size_clean},정상,노출,{qty},{qty}")
        else:
            lines.append(f"정상,노출,{qty},{qty}")
    return "\n".join(lines)


# ---- Image quality helpers ----

def _ensure_min_dimensions(image_bytes):
    """Resize image so both width and height are >= 600px.
    Returns resized image bytes (JPEG format).
    """
    return ensure_marketplace_image_spec(image_bytes)


def _ensure_under_2mb(image_bytes):
    """Backward-compatible helper; now targets 300 KB."""
    return ensure_marketplace_image_spec(image_bytes)


def _ensure_image_meets_spec(image_bytes):
    """Ensure image meets ESM specs: JPEG, >= 600x600, <= 300 KB."""
    return ensure_marketplace_image_spec(image_bytes)


# ---- Phase 2: Concurrent per-batch ----

BASE_REFERENCE_PROMPT = (
    "参考图规则：第一张参考图作为产品主体基准，请保持第一张图中的产品颜色、"
    "型号、款式和关键识别特征一致。不要照搬第一张图的原始构图、背景、"
    "光线和拍摄角度；可以根据用户提示重新设计背景、光线、构图、角度、"
    "质感和商业场景。后续参考图只作为材质、细节和氛围辅助，"
    "不要引入其它颜色、型号或款式。"
)


@dataclass
class UploadImageCandidates:
    main: str
    secondary: list[str]


def normalize_image_url_for_dedupe(url):
    """Normalize source image URLs so resized AliExpress variants dedupe together."""
    if not url or not isinstance(url, str):
        return ""
    raw = url.strip()
    try:
        parts = urlsplit(raw)
    except Exception:
        return raw
    path = re.sub(r"(?i)_thumbnail[^/]*(\.[a-z0-9]+)$", r"\1", parts.path)
    path = re.sub(r"(?i)(\.(?:jpg|jpeg|png|webp|gif))(?:_[^/?#]*)+$", r"\1", path)
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), path, "", ""))


def _dedupe_http_urls(urls, limit=6, exclude_keys=None):
    refs = []
    seen = set(exclude_keys or ())
    for url in urls:
        if not url or not isinstance(url, str) or not url.startswith("http"):
            continue
        key = normalize_image_url_for_dedupe(url)
        if not key or key in seen:
            continue
        seen.add(key)
        refs.append(url)
        if len(refs) >= limit:
            break
    return refs


def _dedupe_exact_http_urls(urls, limit=None):
    refs = []
    seen = set()
    for url in urls:
        if not url or not isinstance(url, str) or not url.startswith("http"):
            continue
        if url in seen:
            continue
        seen.add(url)
        refs.append(url)
        if limit is not None and len(refs) >= limit:
            break
    return refs


def _dedupe_csv_urls(value):
    urls = [u.strip() for u in str(value or "").split(",")]
    return ",".join(_dedupe_exact_http_urls(urls))


def collect_upload_image_candidates(prod, secondary_limit=None):
    """Collect upload image candidates before generation/upload dedupe."""
    if prod.platform == "aliexpress":
        main = prod.variant_imgs[0] if prod.variant_imgs else prod.main_img
        secondary_seed = []
        if prod.variant_imgs:
            secondary_seed.extend(prod.variant_imgs[1:])
        else:
            secondary_seed.extend(prod.extra_imgs)
    else:
        main = prod.main_img
        secondary_seed = list(prod.variant_imgs) + list(prod.extra_imgs)

    main_refs = _dedupe_http_urls([main], limit=1)
    main = main_refs[0] if main_refs else ""
    exclude = {normalize_image_url_for_dedupe(main)} if main else set()
    secondary = _dedupe_http_urls(
        secondary_seed,
        limit=secondary_limit or 10_000,
        exclude_keys=exclude,
    )
    return UploadImageCandidates(main=main, secondary=secondary)


def collect_generation_refs(prod, limit=6):
    """Collect image-generation refs with first image as the base reference."""
    candidates = collect_upload_image_candidates(prod)
    if prod.platform == "aliexpress":
        return _dedupe_http_urls([candidates.main] + candidates.secondary, limit=limit)
    return _dedupe_http_urls([candidates.main] + list(prod.extra_imgs), limit=limit)


def build_generation_prompt(prompt):
    """Prefix prompt with reference-image priority guidance."""
    prompt = (prompt or "").strip()
    if BASE_REFERENCE_PROMPT in prompt:
        return prompt
    return f"{BASE_REFERENCE_PROMPT}\n\n{prompt}" if prompt else BASE_REFERENCE_PROMPT


def _gen_main_image(prod, prompt, storage):
    """Generate main image. AliExpress: 1st variant image as reference. Shein: main+extra."""
    refs = collect_generation_refs(prod)
    prod.ai_source_image_url = refs[0] if refs else ""
    prompt = build_generation_prompt(prompt)
    last_err = None
    for attempt in range(3):  # initial + 2 retries
        try:
            img_bytes = generate_image(prompt, refs)
            img_bytes = _ensure_image_meets_spec(img_bytes)
            url = storage.upload(img_bytes, f"main_{prod.parent_sku}.jpg")
            return url
        except Exception as e:
            last_err = e
            if attempt < 2:
                prod.logs.append(f"{time.strftime('%H:%M:%S')} 生图失败(重试{attempt+1}): {e}")
                time.sleep(3)
    prod.logs.append(f"{time.strftime('%H:%M:%S')} 生图失败(已重试2次): {last_err}")
    return None


def _gen_detail_html(prod, all_products, storage):
    """Generate AB column detail HTML. Thread-safe."""
    try:
        import io
        from PIL import Image
        # AliExpress: use desc_images if available, else fallback to all images+variant
        if prod.platform == "aliexpress":
            if prod.desc_images:
                img_urls = list(prod.desc_images)
            else:
                img_urls = prod.extra_imgs + prod.variant_imgs
        else:
            img_urls = [prod.main_img] + prod.extra_imgs
        for p2 in all_products:
            if p2.parent_sku == prod.parent_sku:
                for u in p2.extra_imgs:
                    if u not in img_urls:
                        img_urls.append(u)
        # Dedup by stem, limit 20
        seen = set()
        deduped = []
        for u in img_urls:
            stem = _url_stem(u)
            if stem not in seen and len(deduped) < 20:
                seen.add(stem)
                deduped.append(u)

        html_parts = []
        for u in deduped:
            try:
                img_data = download_image(u)
                img = Image.open(io.BytesIO(img_data))
                w, h = img.size
                # Resize to width 800; ensure height >= 600 (ESM minimum)
                new_h = max(int(h * 800 / w) if w else 600, 600)
                img = img.resize((800, new_h), Image.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format="JPEG", quality=85)
                # Ensure meets spec before upload
                safe = _ensure_image_meets_spec(buf.getvalue())
                cloud_url = storage.upload(safe, f"detail_{int(time.time())}.jpg")
                html_parts.append(f'<P><img src="{cloud_url}" width="800"></P>')
            except Exception as e:
                prod.logs.append(f"{time.strftime('%H:%M:%S')} 详情图跳过: {e}")
        return "\n".join(html_parts)
    except Exception as e:
        prod.logs.append(f"{time.strftime('%H:%M:%S')} 详情图失败: {e}")
        return ""


def _collect_variant_imgs(prod, storage):
    """Collect variant images for AA column. Downloads, ensures >=600x600 / <=2MB,
    uploads to Cloudinary, returns comma-separated Cloudinary URLs."""
    candidates = collect_upload_image_candidates(prod)
    source_key = normalize_image_url_for_dedupe(prod.ai_source_image_url)
    exclude = {source_key} if source_key else set()
    if prod.platform == "aliexpress":
        urls = _dedupe_http_urls(candidates.secondary, limit=10_000, exclude_keys=exclude)
    else:
        main_key = normalize_image_url_for_dedupe(candidates.main)
        if main_key:
            exclude.add(main_key)
        urls = _dedupe_http_urls(prod.variant_imgs, limit=10_000, exclude_keys=exclude)
        if not urls and len(prod.colors) <= 1:
            urls = _dedupe_http_urls(prod.extra_imgs, limit=2, exclude_keys=exclude)
    # Download, fix, upload each to Cloudinary
    safe_urls = []
    for u in urls:
        try:
            img_data = download_image(u)
            safe = _ensure_image_meets_spec(img_data)
            cloud_url = storage.upload(safe, f"var_{int(time.time())}.jpg")
            safe_urls.append(cloud_url)
        except Exception as e:
            prod.logs.append(f"{time.strftime('%H:%M:%S')} 变种图跳过: {e}")
    return ",".join(_dedupe_exact_http_urls(safe_urls)) if safe_urls else ""


# ---- Main Pipeline ----

def _get_category_zh(kr_path):
    """Translate a Korean category path to Chinese, with JSON cache."""
    if not kr_path:
        return ""
    cache = load_category_zh()
    if kr_path in cache:
        return cache[kr_path]
    # Translate
    try:
        prompt = f"""将以下韩文电商类目路径翻译为中文。只输出中文翻译，不要解释。
韩文：{kr_path}
中文："""
        zh = deepseek_chat(prompt, max_tokens=100, temp=0.3)
        zh = zh.strip()
        cache[kr_path] = zh
        save_category_zh(cache)
        return zh
    except Exception:
        return kr_path  # fall back to Korean


class ProcessingPipeline:
    """Orchestrates the full processing pipeline."""

    def __init__(self, source_path, template_path, output_path, prompt_key, mode_auto,
                 progress_callback=None, stat_callback=None, error_callback=None,
                 store_id="", category_id=""):
        self.source_path = source_path
        self.template_path = template_path
        self.output_path = output_path
        self.prompt_key = prompt_key
        self.mode_auto = mode_auto  # True = full auto, False = error confirm
        self.progress_cb = progress_callback or (lambda n, t, d: None)
        self.stat_cb = stat_callback or (lambda ds, hm, up: None)
        self.error_cb = error_callback or (lambda msg: True)  # returns True to continue
        self.store_id = store_id
        self.category_id = category_id
        self._store_category_cache = {}
        self._stopped = False
        self._pause_event = threading.Event()
        self._pause_event.set()
        self.errors = []

    def stop(self):
        self._stopped = True
        self._pause_event.set()

    def pause(self):
        self._pause_event.clear()

    def resume(self):
        self._pause_event.set()

    def run(self, products):
        """Run the full pipeline on a list of Products."""
        cfg = load_config()
        prompts = load_prompts()
        categories = load_categories(
            os.path.join(SKILL_DIR, "uploader", "카테고리목록(类别代码K列).xls"))
        banned_path = os.path.join(SKILL_DIR, "uploader", "违禁词gmk.txt")
        banned_words = load_banned_words(banned_path) if os.path.exists(banned_path) else []
        storage = create_storage_provider(cfg)
        title_mode = cfg.get("title_mode", "AI重写")
        prompt_text = prompts.get(self.prompt_key, prompts.get("generic", ""))
        store_profile_data = load_store_profiles()
        batch_size = cfg.get("batch_size", 10)
        total = len(products)

        # Init output workbook
        twb, tws, fixed = init_output_workbook(
            self.template_path, self.output_path, cfg, store_id=self.store_id)
        done_skus = detect_completed(self.output_path)
        # Mark already-done products
        for prod in products:
            if prod.parent_sku in done_skus:
                prod.status = ProductStatus.DONE

        pending = [p for p in products if p.status != ProductStatus.DONE]
        start_row = len(done_skus)  # continue from last written row

        # Process in batches
        ds_count = image_count = up_count = 0
        for batch_start in range(0, len(pending), batch_size):
            self._pause_event.wait()
            if self._stopped:
                break

            batch_end = min(batch_start + batch_size, len(pending))
            batch = pending[batch_start:batch_end]
            batch_idx = [products.index(p) for p in batch]

            # Phase 1: Serial
            for prod in batch:
                self._pause_event.wait()
                if self._stopped:
                    break
                real_idx = products.index(prod)
                try:
                    prod.status = ProductStatus.PHASE1_TITLE
                    self.progress_cb(real_idx + 1, total, f"#{real_idx + 1} 标题生成中...")
                    prod.ai_title = phase1_title(prod, banned_words, prompts, title_mode=title_mode)
                    ds_count += 1
                    prod.logs.append(f"{time.strftime('%H:%M:%S')} 标题生成完成")
                    subject_names = _profile_identity_texts(
                        getattr(prod, "subject_profile", {}) or {})
                    if subject_names:
                        prod.logs.append(
                            f"{time.strftime('%H:%M:%S')} 售卖主体: {subject_names[0]}")
                    self.progress_cb(real_idx + 1, total, f"#{real_idx + 1} 标题完成")

                    prod.status = ProductStatus.PHASE1_CATEGORY
                    self.progress_cb(real_idx + 1, total, f"#{real_idx + 1} 类目匹配中...")
                    cat_path, k, l, m = phase1_category(prod, categories)
                    ds_count += int(prod.result.pop("_category_deepseek_calls", 0) or 0)
                    ensure_upload_category_codes(k, l, m)
                    prod.result["K"] = k       # ESM code (number) — what Gmarket expects
                    prod.result["L"] = l       # auction code
                    prod.result["M"] = m       # Gmarket code
                    prod.result["K_path"] = cat_path  # Korean path for display
                    # Translate Korean category path to Chinese (cached)
                    cat_zh = _get_category_zh(cat_path)
                    prod.result["K_zh"] = cat_zh
                    prod.logs.append(f"{time.strftime('%H:%M:%S')} 类目匹配: {cat_path} ({cat_zh})")
                    if self.store_id:
                        cache_key = cat_path or f"{k}|{l}|{m}"
                        store_match = self._store_category_cache.get(cache_key)
                        if store_match is None:
                            store_match = match_store_category(
                                self.store_id, prod, cat_path, cat_zh,
                                store_profile_data, use_ai=True)
                            self._store_category_cache[cache_key] = store_match
                        if store_match.get("method") == "deepseek":
                            ds_count += int(store_match.get("deepseek_calls", 1) or 1)
                        store_category_id = store_match.get("category_id", "")
                        if not store_category_id:
                            evidence = "；".join(store_match.get("evidence", []))
                            raise ValueError(
                                f"店铺类目固定代码匹配失败：{cat_zh or cat_path}。"
                                f"DeepSeek 未从该店铺已录入的大类中返回有效选择。{evidence}")
                        row_fields = category_fields(
                            self.store_id, store_category_id, store_profile_data)
                        if not row_fields.get("AL") or not row_fields.get("AM"):
                            raise ValueError(
                                f"店铺类目 {store_category_id} 未配置完整的 AL/AM")
                        prod.store_id = self.store_id
                        prod.store_category_id = store_category_id
                        prod.logs.append(
                            f"{time.strftime('%H:%M:%S')} 店铺固定代码类目: "
                            f"{store_category_id} (AL={row_fields['AL']}, AM={row_fields['AM']})")
                    self.progress_cb(real_idx + 1, total, f"#{real_idx + 1} 类目完成")

                    prod.status = ProductStatus.PHASE1_PRICE
                    price = phase1_price(prod, products, cfg)
                    prod.result["O"] = price
                    prod.result["P"] = price
                    prod.logs.append(f"{time.strftime('%H:%M:%S')} 价格计算: {price}")
                    self.progress_cb(real_idx + 1, total, f"#{real_idx + 1} Phase1完成")

                    prod.status = ProductStatus.PHASE1_DONE
                    prod.result["W"] = phase1_w_type(prod)
                    prod.result["X"] = phase1_x_attr(prod, prod.result["W"])
                    prod.result["Y"] = phase1_y_list(prod, prod.result["W"], prod.result["X"])
                except Exception as e:
                    prod.status = ProductStatus.FAILED
                    prod.logs.append(f"{time.strftime('%H:%M:%S')} Phase1 失败: {e}")
                    self.errors.append((products.index(prod), str(e)))

            # Phase 2: Concurrent (thread pool)
            if self._stopped:
                break

            z_results = {}
            ab_results = {}
            aa_results = {}

            def worker(prod):
                idx = products.index(prod)
                prod.status = ProductStatus.PHASE2_MAIN_IMG
                self.progress_cb(idx + 1, total, f"#{idx + 1} 生图中...")
                main_url = _gen_main_image(prod, prompt_text, storage)
                if main_url is None:
                    prod.status = ProductStatus.FAILED
                    z_results[idx] = ""
                    ab_results[idx] = ""
                    aa_results[idx] = ""
                    return 0, 0
                z_results[idx] = main_url

                prod.status = ProductStatus.PHASE2_DETAIL
                self.progress_cb(idx + 1, total, f"#{idx + 1} 详情图制作中...")
                ab_results[idx] = _gen_detail_html(prod, products, storage)
                detail_count = ab_results[idx].count("<P>") if ab_results[idx] else 1

                prod.status = ProductStatus.PHASE2_VARIANT
                aa_results[idx] = _collect_variant_imgs(prod, storage)
                self.progress_cb(idx + 1, total, f"#{idx + 1} 生成完成")

                return 1, max(detail_count, 1)  # (image_count, up_count)

            threads = []
            thread_results = []  # collect (image_count, up_count) from each worker
            ready_batch = [prod for prod in batch if prod.status != ProductStatus.FAILED]
            for prod in ready_batch:
                def _runner(p=prod):
                    r = worker(p)
                    thread_results.append(r)
                t = threading.Thread(target=_runner)
                threads.append(t)
                t.start()

            for t in threads:
                t.join()
            for r in thread_results:
                image_count += r[0]
                up_count += r[1]

            # Write batch
            for prod in batch:
                idx = products.index(prod)
                if prod.status == ProductStatus.FAILED:
                    self.progress_cb(idx + 1, total, f"#{idx + 1} 失败，未导出")
                    continue
                prod.result["Z"] = z_results.get(idx, "")
                prod.result["AA"] = aa_results.get(idx, "")
                prod.result["AB"] = ab_results.get(idx, "")
                prod.status = ProductStatus.DONE
                selected_category = getattr(prod, "store_category_id", "") or self.category_id
                row_category_fields = category_fields(
                    self.store_id, selected_category, store_profile_data)
                write_product_row(tws, start_row, prod, fixed, 8, row_category_fields)
                start_row += 1
                self.progress_cb(idx + 1, total, f"#{idx + 1} 已完成")

            twb.save(self.output_path)
            self.progress_cb(batch_end, total, f"批次 {batch_start//batch_size + 1} 保存完成")
            self.stat_cb(ds_count, image_count, up_count)

            # Error confirm mode
            batch_errors = [e for e in self.errors if batch_start <= e[0] < batch_end]
            if batch_errors and not self.mode_auto:
                msg = f"本批 {len(batch_errors)} 个产品失败:\n"
                msg += "\n".join([f"  #{e[0] + 1}: {e[1][:60]}" for e in batch_errors[:5]])
                self._pause_event.clear()  # pause for user decision
                continue_ok = self.error_cb(msg)
                self._pause_event.set()
                if not continue_ok:
                    self._stopped = True
                    break

        twb.save(self.output_path)
        return ds_count, image_count, up_count
